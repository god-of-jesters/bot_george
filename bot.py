import asyncio
from email import message
import os
import csv
import io
from datetime import datetime
from random import choice
from io import BytesIO
import aiosqlite
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from aiogram.types import BufferedInputFile

import aiosqlite
from entityes.promokod import Promokod
from repo.promokod_repo import add_promokod, get_promo_by_pharse, update_promokod
from dotenv import load_dotenv
from collections import defaultdict

from aiogram import Bot, Dispatcher, Router, F
from aiogram.types import Message, CallbackQuery, callback_query
from aiogram.filters import CommandStart, Command
from aiogram.fsm.storage.memory import MemoryStorage
from aiogram.fsm.context import FSMContext
from aiogram.types import FSInputFile

from database import load_datastore, USERS, TEAMS, FILES, COMPLAINTS, PRODUCT_NAME_INDEX, PRODUCTS
from entityes import product, user
from repo.team_repo import *
from repo.user_repo import *
from repo.file_repo import *
from repo.complaint_repo import *
from repo.product_repo import *
from repo.message_repo import Message as ms, update_status, update_status_skip_new
from repo.message_repo import add_message, delete_message, update_message, get_message, get_new_messages, get_message_access
from entityes.sequence import *
from entityes.logger import *

from keyboards import *
import logging

logging.basicConfig(level=logging.INFO)

load_dotenv()
mode = 'test' if os.getenv("MODE") == "test" else 'production'
API_TOKEN = os.getenv("BOT")
active_sessions = {}
registration = {}
complaintes = {}
complaints_adresat = {}
alarm : dict[int: dict[int, int]] = {}
process_al = {}
al = {}
edit_users = {}
special_step = {}
maling = {}
maling_special = {}
rating = {}
rating_choice = {}
promos = {}
messagess = {}
violetion_vines = {
    "Нахождение на базе без личной карточки участника": 10,
    "Нахождение в неподходящей под погодные условия одежде на улице": 7,
    "Употребление никотиносодержащей продукции в неположенном месте": 15,
    "Нецензурная речь": 15,
    "Провоз и употребление энергетических напитков": 10,
    "Нахождение после отбоя участника вне своей комнаты": 20,
    "Нарушение общественного спокойствия после отбоя": 25,
    "Пропуск программных моментов без уважительной причины": 20,
    "Порча имущества базы/организаторов/участников": 30,
    "Кража": 50,
    "Оскорбления/конфликты на почве розни": 50,
    "Опьянение": 50,
}
_album_buffer: dict[tuple[int, str], list[Message]] = defaultdict(list)
_album_tasks: dict[tuple[int, str], asyncio.Task] = {}

MAX_PHOTOS = 3
MAX_VIDEOS = 1
ALBUM_FLUSH_DELAY = 0.7
router = Router()

@router.callback_query(DecisionCb.filter())
async def on_decision(call: CallbackQuery, callback_data: DecisionCb):
    req_id = callback_data.req_id
    action = callback_data.action  # "ok" / "no"

    async with aiosqlite.connect(DB_PATH) as db:
        await db.execute("PRAGMA foreign_keys=ON;")
        await db.execute("BEGIN IMMEDIATE;")
        db.row_factory = aiosqlite.Row

        cur = await db.execute(
            "SELECT id, target_badge, points, status FROM requests WHERE id = ?",
            (req_id,)
        )
        row = await cur.fetchone()

        if not row:
            await db.execute("ROLLBACK;")
            await call.answer("Уведомление не найдено", show_alert=True)
            return

        if row["status"] != "new":
            await db.execute("ROLLBACK;")
            await call.answer("Уже обработано", show_alert=True)
            return

        new_status = "approved" if action == "ok" else "rejected"

        await db.execute(
            "UPDATE requests SET status = ? WHERE id = ?",
            (new_status, req_id)
        )

        if action == "ok":
            await db.execute(
                "UPDATE users SET reiting = reiting - ? WHERE badge_number = ?",
                (row["points"], row["target_badge"])
            )

        await db.commit()

    await call.message.edit_reply_markup(reply_markup=None)

    if action == "ok":
        await call.answer("Принято – очки списаны")
    else:
        await call.answer("Отклонено")

    await show_main_menu(call.bot, call.from_user.id)

async def _apply_complaint_decision(bot: Bot, reviewer_id: int, com: "Complaint", decision: str):
    adr_user = await get_user_by_badge(com.adresat)
    fine = violetion_vines.get(com.violetion, 0)

    if decision == "yes":
        if adr_user and adr_user.role == "Участник":
            await subtract_rating(adr_user.badge_number, fine)
            if adr_user.tg_id is not None:
                await bot.send_message(
                    adr_user.tg_id,
                    f"Жалоба на вас обработана.\n"
                    f"Время жалобы: {com.date_created}.\n"
                    f"Описание: {com.description}"
                )
            com.execution = "done"
    else:
        if adr_user and adr_user.tg_id is not None:
            await bot.send_message(
                adr_user.tg_id,  
                "На вас была подана жалоба. Отдель справедливости посчитала, что жалоба недействительна."
            )
        com.execution = "rejected"

    await update_execution(com.complaint_id, 'done')

    # Notify the user who filed the complaint after it was processed
    verdict_text = "удовлетворена" if decision == "yes" else "не удовлетворена"
    try:
        await bot.send_message(
            com.user_id,
            "Ваша жалоба обработана.\n"
            f"Нарушение: {getattr(com, 'violetion', '—')}\n"
            f"Описание: {getattr(com, 'description', '—')}\n"
            f"Решение: {verdict_text}"
        )
    except Exception:
        # If user blocked the bot / chat unavailable, don't break the reviewer flow
        pass

@router.callback_query(RoomComplaintCb.filter(), MainMenu.manage_rooms)
async def process_manage_rooms_callback(callback_query: CallbackQuery, callback_data: RoomComplaintCb, state: FSMContext,):
    user_id = callback_query.from_user.id
    action = callback_data.action
    complaint_id = callback_data.complaint_id

    execution = "done" if action == "agree" else "rejected"

    await update_execution(complaint_id, execution)

    complaint = await get_complaint(complaint_id)
    if complaint:
        await notify_persone_room_problems(callback_query.bot, complaint, action, state)
    else:
        await show_main_menu(callback_data.bot, user_id, state)
        return

    await callback_query.message.edit_reply_markup(reply_markup=None)
    await callback_query.answer("Готово")

    if action == "agree":
        await callback_query.message.answer("Жалоба обработана.")
    else:
        await callback_query.message.answer("Жалоба отклонена.")

    await show_next_room_problem(callback_query.bot, user_id, state)

@router.callback_query(RoomComplaintCb.filter())
async def process_complaint_from_main(callback_query: CallbackQuery, callback_data: RoomComplaintCb, state: FSMContext):
    user_id = callback_query.from_user.id
    action = callback_data.action
    complaint_id = callback_data.complaint_id

    if action == "agree":
        await callback_query.message.answer("Жалоба обработана.")
        execution = "done"
    else:
        await callback_query.message.answer("Жалоба отклонена.")
        execution = "done"  # или "done", как у тебя принято

    await update_execution(complaint_id, execution)

    complaint = await get_complaint(complaint_id)
    if complaint:
        await notify_persone_room_problems(callback_query.bot, complaint, action, state)

    await callback_query.message.edit_reply_markup(reply_markup=None)
    await show_main_menu(callback_query.bot, user_id, state)
    await state.set_state(MainMenu.main_menu_admins)

@router.callback_query(lambda c: c.data in ("yes", "no"), ComplaintReview.main)
async def process_complaint_from_main(callback_query: CallbackQuery, state: FSMContext):
    await callback_query.answer()
    user_id = callback_query.from_user.id

    com = al.get(user_id)
    if not com:
        await callback_query.message.answer("Текущая жалоба не найдена.")
        await show_main_menu(callback_query.bot, user_id, state)
        return

    await _apply_complaint_decision(callback_query.bot, user_id, com, callback_query.data)
    al.pop(user_id, None)

    if callback_query.data == "yes":
        await callback_query.answer("Успешно сняли рейтинг")
        complaint = await get_oldest_complaint()
        if complaint:
            al[user_id] = complaint
            await send_complaint_files(callback_query.bot, user_id, complaint.complaint_id)
            adr = await get_user_by_badge(complaint.adresat) if complaint.adresat else None
            adr_str = adr.fio if adr else f"бейдж {complaint.adresat}" if complaint.adresat else "—"
            await callback_query.message.answer(
                f"Статус: {complaint.status}\nДата создания: {complaint.date_created}\nНа: {adr_str}\nЖалоба: {complaint.description}\nТип жалобы: {complaint.violetion}",
                reply_markup=get_yes_no_keyboard()
            )
            await state.set_state(ComplaintReview.main)
        else:
            await show_main_menu(callback_query.bot, user_id)
    else:
        await callback_query.answer("Успешно защитили человека")
        complaint = await get_oldest_complaint()
        if complaint:
            al[user_id] = complaint
            await send_complaint_files(callback_query.bot, user_id, complaint.complaint_id)
            adr = await get_user_by_badge(complaint.adresat) if complaint.adresat else None
            adr_str = adr.fio if adr else f"бейдж {complaint.adresat}" if complaint.adresat else "—"
            await callback_query.message.answer(
                f"Статус: {complaint.status}\nДата создания: {complaint.date_created}\nНа: {adr_str}\nЖалоба: {complaint.description}\nТип жалобы: {complaint.violetion}",
                reply_markup=get_yes_no_keyboard()
            )
            await state.set_state(ComplaintReview.main)
        else:
            await show_main_menu(callback_query.bot, user_id)

@router.callback_query(lambda c: c.data in ("yes", "no"), ComplaintReview.stat)
async def process_complaint_fate(callback_query: CallbackQuery, state: FSMContext):
    await callback_query.answer()
    user_id = callback_query.from_user.id

    com = al.get(user_id)
    if not com:
        await callback_query.message.answer("Текущая жалоба не найдена.")
        await show_main_menu(callback_query.bot, user_id, state)
        return

    await _apply_complaint_decision(callback_query.bot, user_id, com, callback_query.data)
    al.pop(user_id, None)

    other = await get_oldest_complaint()
    if other and other.status == "alert":
        await callback_query.message.answer(
            "Жалоба успешно обработана. Есть еще срочные жалобы, ответить?",
            reply_markup=get_yes_no_keyboard()
        )
        await state.set_state(ComplaintReview.safe)
        return

    await callback_query.message.answer("Жалоба успешно обработана.")
    await show_main_menu(callback_query.bot, user_id, state)

@router.callback_query(lambda c: c.data in ("yes", "no"))
async def process_alarm_complaint(callback_query: CallbackQuery, state: FSMContext):
    await callback_query.answer()
    user_id = callback_query.from_user.id

    if callback_query.data == "no":
        await callback_query.message.answer("Скип скип")
        await show_main_menu(callback_query.bot, user_id, state)
        return

    queue = alarm.get(user_id, [])
    if not queue:
        await callback_query.message.answer("Срочных жалоб сейчас нет.")
        await show_main_menu(callback_query.bot, user_id, state)
        return

    com = None
    while queue:
        cid = queue[0]
        cand = await get_complaint(cid)

        if cand and cand.execution == "new" and cand.status == "alert":
            com = cand
            queue.pop(0)
            break
        else:
            queue.pop(0)

    if not com:
        await callback_query.message.answer("Срочных жалоб сейчас нет.")
        await show_main_menu(callback_query.bot, user_id, state)
        return

    process_al.setdefault(user_id, []).append(com.complaint_id)
    await update_execution(com.complaint_id, "view")

    user = await get_user(com.user_id)
    adr = await get_user_by_badge(com.adresat)
    adr_name = adr.fio if adr else f"бейдж {com.adresat}"

    await send_complaint_files(callback_query.bot, user_id, com.complaint_id)
    await callback_query.message.answer(
        f"Жалоба от {user.fio}\nНа {adr_name}\nКатегория жалобы: {com.violetion}\nЖалоба: {com.description}",
        reply_markup=get_yes_no_keyboard()
    )

    al[user_id] = com
    await state.set_state(ComplaintReview.stat)

@router.message(MainMenu.profile)
async def main_menu_callback(message: Message, state: FSMContext):
    user_id = message.from_user.id
    match active_sessions[user_id].role:
            case "Участник":
                await message.answer("Главное меню.", reply_markup=get_main_menu_student_keyboard())
                await state.set_state(MainMenu.main_menu_student)
            case "Организатор":
                await message.answer("Главное меню.", reply_markup=get_main_menu_organizer_keyboard())
                await state.set_state(MainMenu.main_menu_organizer)
            case "РПГ":
                await message.answer("Главное меню.", reply_markup=get_main_menu_rpg_organizer_keyboard())
                await state.set_state(MainMenu.main_menu_rpg_organizer) 
            case "Администраторы по комнатам":
                await message.answer("Главное меню.", reply_markup=get_main_menu_admins_keyboard())
                await state.set_state(MainMenu.main_menu_admins)
            case "Рейтинг":
                await message.answer("Главное меню.", reply_markup=get_main_menu_rating_team_keyboard())
                await state.set_state(MainMenu.main_menu_rating_team) 
            case "Медиа":
                await message.answer("Главное меню.", reply_markup=get_main_menu_media_team_keyboard())
                await state.set_state(MainMenu.main_menu_media)
            case "Главный организатор":
                await message.answer("Главное меню.", reply_markup=get_main_menu_chief_organizer_keyboard())
                await state.set_state(MainMenu.main_menu_chief_organizer)

@router.message(Command("teg"))
async def cmd_teg(message: Message):
    id = message.from_user.id
    await message.answer(f"{id}")

@router.message(Command('exit'))
async def cmd_exit(message: Message):
    await message.answer("Вы вышли из бота. Для регистрации введите команду /start")
    active_sessions.pop(message.from_user.id)
    await del_from_active(message.from_user.id)

"""MAIN MENUS"""

@router.callback_query(MainMenu.main_menu_student)
async def show_profile(callback_query: CallbackQuery, state: FSMContext):
    user_id = callback_query.from_user.id
    data = callback_query.data
    match data:
        case "profile":
            profile = "Профиль:\n"
            profile += f"ФИО: {active_sessions[user_id].fio}\n Роль: {active_sessions[user_id].role}\n"
            profile += f"Номер бейджа: {active_sessions[user_id].badge_number}\n"
            if active_sessions[user_id].role == "Участник":
                profile += f"Название команды: {TEAMS[active_sessions[user_id].team_number].team_name}\n" if active_sessions[user_id].team_number in TEAMS else "Не назначена команда\n"
                profile += f"Рейтинг: {active_sessions[user_id].reiting}\n"
                profile += f"Рейтинг команды: {TEAMS[active_sessions[user_id].team_number].reiting}\n" if active_sessions[user_id].team_number in TEAMS else ""
                profile += f"Баланс: {active_sessions[user_id].balance} очков\n"
                await callback_query.message.answer(profile, reply_markup=get_profile_keyboard())
                await state.set_state(MainMenu.profile)
        case "complaint":
            await callback_query.message.answer("На что будет жалоба.", reply_markup=get_complaint_keyboard())
            await state.set_state(MainMenu.complaint)
        case "my_complaints":
            s = ''
            complaints = await get_user_complaints(user_id)
            for i, com in enumerate(complaints):
                adr = await get_user_by_badge(com.adresat) if com.adresat else None
                adr_name = adr.fio if adr else f"бейдж {com.adresat}" if com.adresat else "—"
                s += 'Номер жалобы:' + str(i + 1) + '\n'
                s += 'На: ' + adr_name + '\n'
                s += 'Описание: ' + com.description + '\n\n'
            await callback_query.message.answer(s or "У вас пока нет жалоб.", reply_markup=get_profile_keyboard())
        case "entertainment":
            await state.set_state(MainMenu.student_entertainment)
            await callback_query.message.answer("Развлечения.", reply_markup=get_student_entertainment_keyboard())
        case "help":
            await callback_query.message.answer("Помощь.", reply_markup=get_student_help_keyboard())
            await state.set_state(MainMenu.student_help)
        case "message_to_admin":
            access = await get_message_access(user_id)
            if access:
                await callback_query.message.answer("Напишите ваше сообщение администрации.")
                await state.set_state(MainMenu.message_to_admin)
            else:
                await callback_query.message.answer("Сообщения можно отправлять только раз в полчаса.")
                await state.set_state(MainMenu.main_menu_student)
        case _:
            await callback_query.message.answer("Команда не распознана.")

@router.callback_query(MainMenu.main_menu_organizer)
async def show_main_organizer(callback_query: CallbackQuery, state: FSMContext):
    user_id = callback_query.from_user.id
    data = callback_query.data
    role = getattr(active_sessions.get(user_id), "role", None)
    if role != "Организатор":
        return

    match data:
        case "profile":
            await callback_query.message.answer("Профиль организатора.", reply_markup=get_profile_keyboard())
            await state.set_state(MainMenu.profile)

        case "complaint":
            await callback_query.message.answer("Подать жалобу.", reply_markup=get_complaint_keyboard())
            await state.set_state(MainMenu.complaint)

        case "view_complaints":
            complaints = await get_user_complaints(user_id)
            if complaints:
                await callback_query.message.answer("Жалобы в работе.")
                for complaint in complaints:
                    author = await get_user(complaint.user_id)
                    adr = await get_user_by_badge(complaint.adresat) if complaint.adresat else None
                    adr_str = adr.fio if adr else f"бейдж {complaint.adresat}" if complaint.adresat else "—"
                    await callback_query.message.answer(
                        f"Статус: {complaint.status}\nДата: {complaint.date_created}\nНа: {adr_str}\nЖалоба: {complaint.description}\n"
                    )
            else:
                await callback_query.message.answer("Жалобы в работе не найдены.")
                await show_main_menu(callback_query.bot, user_id, state)
            

        case "contact":
            await callback_query.message.answer("Напишите сообщение для команды рейтинга.")
            await state.set_state(MainMenu.message_to_rating_team)

        case "mailing":
            await callback_query.message.answer("Введите текст рассылки.\n")
            await state.set_state(Mailing.waiting_for_mailing_text)

        case "help":
            await callback_query.message.answer("Помощь.\n(Здесь будет справка для организаторов.)")

        case _:
            await callback_query.message.answer("Команда не распознана.")

@router.callback_query(MainMenu.main_menu_rpg_organizer)
async def show_main_rpg_organizer(callback_query: CallbackQuery, state: FSMContext):
    user_id = callback_query.from_user.id
    data = callback_query.data
    role = getattr(active_sessions.get(user_id), "role", None)
    if role != "РПГ":
        return

    match data:
        case "profile":
            await callback_query.message.answer("Профиль РПГ.", reply_markup=get_profile_keyboard())
            await state.set_state(MainMenu.profile)
        
        case "create_promo":
            await callback_query.message.answer("Введите фразу промокода")
            await state.set_state(PromoCreate.waiting_for_phrase)

        case "bonus":
            await callback_query.message.answer("Выберите получателя бонуса", reply_markup=get_maling_adresat())
            await state.set_state(Bonus.waiting_adresat)

        case "edit_products":
            await callback_query.message.answer('Выберете дейстивие', reply_markup=get_edit_product_choice())
            await state.set_state(Products.wait_choice_action)

        case "get_sells":
            await export_sells_xlsx(callback_query.bot, user_id)
        
        case "get_products":
            await export_remaining_products_xlsx(callback_query.bot, user_id)

        case "complaint":
            await callback_query.message.answer("На что будет жалоба.", reply_markup=get_complaint_keyboard())
            await state.set_state(MainMenu.complaint)

        case "m":
            await callback_query.message.answer("Выберете получателей рассылки.\n", reply_markup=get_maling_adresat())
            await state.set_state(Mailing.waiting_adresat)

        case "contact":
            await callback_query.message.answer("Напишите сообщение для команды рейтинга.")
            await state.set_state(MainMenu.message_to_rating_team)

        case "help":
            await callback_query.message.answer("Помощь.\n(Здесь будет справка для РПГ-организаторов.)")

        case _:
            await callback_query.message.answer("Команда не распознана.")

@router.callback_query(MainMenu.main_menu_admins)
async def show_main_admins(callback_query: Message, state: FSMContext):
    user_id = callback_query.from_user.id
    data = callback_query.data
    role = getattr(active_sessions.get(user_id), "role", None)
    if role != "Администратор":
        return

    match data:
        case "profile":
            await callback_query.message.answer("Профиль администратора по комнатам.", reply_markup=get_profile_keyboard())
            await state.set_state(MainMenu.profile)

        case "manage_rooms":
            await show_next_room_problem(callback_query.bot, user_id, state)

        case "mailing":
            await callback_query.message.answer("Выберете получателей рассылки.\n", reply_markup=get_maling_adresat())
            await state.set_state(Mailing.waiting_adresat)
        
        case 'complaint':
            await callback_query.message.answer("На что будет жалоба.", reply_markup=get_complaint_keyboard())
            await state.set_state(MainMenu.complaint)

        case "contact":
            await callback_query.message.answer("Напишите сообщение для команды рейтинга.")
            await state.set_state(MainMenu.message_to_rating_team)

        case "help":
            await callback_query.message.answer("Помощь.\n(Здесь будет справка для администраторов по комнатам.)")

        case _:
            await callback_query.message.answer("Команда не распознана.")

@router.callback_query(MainMenu.main_menu_rating_team)
async def show_main_rating_team(callback_query: CallbackQuery, state: FSMContext):
    user_id = callback_query.from_user.id
    data = callback_query.data
    role = getattr(active_sessions.get(user_id), "role", None)
    if role != "Рейтинг":
        return

    match data:
        case "profile":
            await callback_query.message.answer("Профиль команды рейтинга.", reply_markup=get_profile_keyboard())
            await state.set_state(MainMenu.profile)

        case "view_complaints":
            complaint = await get_oldest_complaint()
            if complaint:
                al[user_id] = complaint
                await send_complaint_files(callback_query.bot, user_id, complaint.complaint_id)
                adr = await get_user_by_badge(complaint.adresat) if complaint.adresat else None
                adr_str = adr.fio if adr else f"бейдж {complaint.adresat}" if complaint.adresat else "—"
                await callback_query.message.answer(
                    f"Статус: {complaint.status}\nДата создания: {complaint.date_created}\nНа: {adr_str}\nЖалоба: {complaint.description}\nТип жалобы: {complaint.violetion}",
                    reply_markup=get_yes_no_keyboard()
                )
                await state.set_state(ComplaintReview.main)
            else:
                await callback_query.message.answer(f'Пока что нет жалоб.')
                await state.set_state(MainMenu.main_menu_rating_team)
                await show_main_menu(callback_query.bot, user_id, state)

        case "participants":
            mes = await get_roles_stats_message()
            await callback_query.message.answer(mes, reply_markup=get_users_keyboard())
            await state.set_state(MainMenu.users)

        case "assign_rating":
            await callback_query.message.answer("Начисление и штрафы. Введите бейдж участника\n")
            await state.set_state(Rating.waiting_for_badge_number)
        
        case 'bonus':
            await callback_query.message.answer("Введите номер бейджа\n")
            await state.set_state(Rating.waiting_for_badge_number_bonus)

        case "inbox_messages":
            await update_status_skip_new()
            messages = await get_new_messages()
            await callback_query.message.answer(f"Входящие сообщения: {len(messages)}\n")
            if messages:
                m = messages[0]
                messagess[user_id] = m
                us = await get_user(m.user_id)
                await callback_query.bot.send_message(user_id, f'Сообщение от {us.fio} {us.username}\n\n' + m.text, reply_markup=get_message_keyboard())
            await state.set_state(YesNoChoice.waiting_for_message_answer)

        case "mailing":
            await callback_query.message.answer("Выберете получателей рассылки.\n", reply_markup=get_maling_adresat())
            await state.set_state(Mailing.waiting_adresat)
        
        case "complaint":
            await callback_query.message.answer("На что будет жалоба.", reply_markup=get_complaint_keyboard())
            await state.set_state(MainMenu.complaint)

        case "help":
            await callback_query.message.answer("Помощь.\n(Здесь будет справка для команды рейтинга.)")

        case _:
            await callback_query.message.answer("Команда не распознана.")

@router.callback_query(MainMenu.main_menu_media)
async def show_main_chief_organizer(callback_query: Message, state: FSMContext):
    user_id = callback_query.from_user.id
    data = callback_query.data
    role = getattr(active_sessions.get(user_id), "role", None)
    if role != "Медиа":
        return

    match data:
        case "profile":
            await callback_query.message.answer("Профиль медиа.", reply_markup=get_profile_keyboard())
            await state.set_state(MainMenu.profile)

        case "mailing":
            await callback_query.message.answer("Выберете получателей рассылки.\n", reply_markup=get_maling_adresat())
            await state.set_state(Mailing.waiting_adresat)

        case "contact":
            await callback_query.message.answer("Напишите сообщение для команды рейтинга.")
            await state.set_state(MainMenu.message_to_rating_team)

        case "complaint":
            await callback_query.message.answer("На что будет жалоба.", reply_markup=get_complaint_keyboard())
            await state.set_state(MainMenu.complaint)

        case "help":
            await callback_query.message.answer("Помощь.\n(Здесь будет справка для медиа.)")

        case _:
            await callback_query.message.answer("Команда не распознана.")

"""STUDENT MENU"""

@router.callback_query(MainMenu.student_entertainment)
async def show_student_entertainment(callback_query: CallbackQuery, state: FSMContext):
    user_id = callback_query.from_user.id
    data = callback_query.data
    role = getattr(active_sessions.get(user_id), "role", None)
    if role != "Участник":
        return
    match data:
        case "shop":
            await callback_query.message.answer("Магазин.", reply_markup=get_student_shop_keyboard())
            await state.set_state(Shop.wait_for_choice_action)
        case "zags":
            await callback_query.message.answer("ЗАГС.", reply_markup=get_student_zags_keyboard())
            await state.set_state(ZAGS.waiting_for_choice)
        case "back_to_main_menu":
            await show_main_menu(callback_query.bot, user_id, state)
        case _:
            await callback_query.message.answer("Команда не распознана.")

@router.callback_query(MainMenu.student_help)
async def show_student_help(callback_query: CallbackQuery, state: FSMContext):
    user_id = callback_query.from_user.id
    data = callback_query.data
    role = getattr(active_sessions.get(user_id), "role", None)
    if role != "Участник":
        return
    match data:
        case "rules":
            await callback_query.message.answer("Правила и обязанности участника.")
        case "help_in_work":
            await callback_query.message.answer("Помощь по работе с ботом.")
        case _:
            await callback_query.message.answer("Команда не распознана.")
    await show_main_menu(callback_query.bot, user_id, state)

@router.message(MainMenu.message_to_admin)
async def process_message_to_admin(message: Message, state: FSMContext):
    user_id = message.from_user.id
    message_text = message.text
    await add_message(
        ms(
            user_id=user_id,
            adresat="Рейтинг",
            badge_number=0,
            text=message_text,
        )
    )
    await message.answer("Ваше сообщение отправлено администрации. Спасибо.")
    await show_main_menu(message.bot, user_id, state)

@router.message(MainMenu.message_to_rating_team)
async def process_message_to_rating_team(message: Message, state: FSMContext):
    user_id = message.from_user.id
    role = getattr(active_sessions.get(user_id), "role", None)
    badge = active_sessions[user_id].badge_number
    if role != "Администратор" and role != "Медиа":
        return
    match role:
        case "Администратор":
            adresat = "Администраторы по комнатам"
        case "Медиа":
            adresat = "Медиа"
    await add_message(
        ms(
            user_id=user_id,
            adresat=adresat,
            badge_number=badge,
            text=message.text,
        )
    )
    await message.answer("Ваше сообщение отправлено команде рейтинга. Спасибо.")
    await show_main_menu(message.bot, user_id, state)

"""COMPLAINTS"""

@router.callback_query(lambda c: c.data == "complaint_done", ComplaintProcess.waiting_for_complaint_files)
async def finish_complaint_cb(callback: CallbackQuery, state: FSMContext):
    user_id = callback.from_user.id
    await callback.answer()
    complaint = complaintes[user_id]
    complaint_id = await add_complaint(complaint)
    await log_complaint_created(complaint.user_id, complaint.adresat, complaint_id)

    if complaint.status == "alert":
        if user.role != 'Участник':
            await send_complaint_notify(message.bot, complaint, state)
        else:
            await notify_all_reiting_team(message.bot, complaint, state)
    if complaint.status == 'soon':
        await send_complaint_notify_soon(callback.bot, complaint)
    if complaint.status == 'room_problems':
        await send_complaint_room_problems(message.bot, complaint)
    if complaint.status == 'other':
        await send_complaint_room_problems(message.bot, complaint)

    if user_id in complaintes:
        del complaintes[user_id]

    await callback.message.answer("Жалоба отправлена. Спасибо.")
    await show_main_menu(callback.bot, user_id, state)

@router.callback_query(MainMenu.complaint)
async def process_complaint_callback(callback_query: Message, state: FSMContext):
    data = callback_query.data
    user_id = callback_query.from_user.id
    roles = {
        "participant_behavior": "Участник",
        "organizer_behavior": "Организатор",
        "room_problems": "Комнатные проблемы",
        "other": "Другое"
    }
    if user_id not in complaintes:
        complaintes[user_id] = Complaint(user_id=user_id, status="Новая")
        complaints_adresat[user_id] = roles[data]
    if roles[data] != 'Участник' and roles[data] != 'Организатор':
        complaintes[user_id].status = data
        complaintes[user_id].adresat = 0
        await callback_query.message.answer("Введите текст жалобы, если проблемы с комнатой то укажите в тексте её номер, если что-то другое, то как можно подробнее опишите проблему.")
        await state.set_state(ComplaintProcess.waiting_for_complaint_text)
    else:
        await callback_query.message.answer("Введите номер бейджа.")
        await state.set_state(ComplaintProcess.waiting_for_badge)

@router.message(ComplaintProcess.waiting_for_badge)
async def process_complaint_badge(message: Message, state: FSMContext):
    user_id = message.from_user.id
    if not message.text.isdigit():
        await message.answer("Номер должен быть числом, введите еще раз")
        return
    num = int(message.text)
    user = await get_user_by_badge(num)
    if user:
        if num == active_sessions[user_id].badge_number:
            await message.answer('Нельзя подать жалобу на себя же. Введите еще раз номер бейджа')
            await state.set_state(ComplaintProcess.waiting_for_badge)
            return
        if user.role != 'Участник' and user.badge_number >= 100:
            await message.answer('Неверно введет номер бейджа для организатора, введите еще раз')
            await state.set_state(ComplaintProcess.waiting_for_badge)
            return
        if user.role == 'Участник' and user.badge_number < 100:
            await message.answer('Неверно введет номер бейджа для участника, введите еще раз')
            await state.set_state(ComplaintProcess.waiting_for_badge)
            return
        complaintes[user_id].adresat = user.badge_number
        
        await message.answer(text="Выберете категорию жалобы.", reply_markup=get_complaint_category_keyboard())
        await state.set_state(ComplaintProcess.waiting_for_complaint_category)
    else:
        await message.answer('Такого человека не существует, введите номер бейджа еще раз')
        return

@router.callback_query(ComplaintProcess.waiting_for_complaint_category)
async def process_complaint_category_callback(callback_query: Message, state: FSMContext):
    text_alert = "Выберете тип нарушения \n\
    1. Провоз и употребление энергитических напитков \n\
    2. Нахождение после отбоя участника вне своей комнаты \n\
    3. Нарушение общественного спокойствия после отбоя \n\
    4. Порча имущества базы/организаторов/участников \n\
    5. Кража \n\
    6. Опьянение"

    text_soon = 'Выберете тип нарушения \n\
    1. Нахождение на базе без личной карточки участника \n\
    2. Нахождение в неподходящей под погодные условия одежде на улице \n\
    3. Употребление никотиносодержащей продукции в неположенном месте \n\
    4. Нецензурная речь\n'

    data = callback_query.data
    user_id = callback_query.from_user.id
    if user_id in complaintes:
        complaintes[user_id].status = data
        match data:
            case 'alert':
                await callback_query.bot.send_message(user_id, text_alert, reply_markup=get_alert_keyboard())
            case 'soon':
                await callback_query.bot.send_message(user_id, text_soon, reply_markup=get_soon_keyboard())
        await state.set_state(ComplaintProcess.waiting_for_violation_type)

@router.callback_query(ComplaintProcess.waiting_for_violation_type)
async def process_complaint_violation_type(callback_query: CallbackQuery, state: FSMContext):
    data = int(callback_query.data)
    user_id = callback_query.from_user.id
    c1 = {
        1: 'Провоз и употребление энергитических напитков',
        2: 'Нахождение после отбоя участника вне своей комнаты',
        3: 'Нарушение общественного спокойствия после отбоя',
        4: 'Порча имущества базы/организаторов/участников',
        5: 'Кража',
        6: 'Опьянение'
    }
    c2 = {
        1: 'Нахождение на базе без личной карточки участника',
        2: 'Нахождение в неподходящей под погодные условия одежде на улице',
        3: 'Употребление никотиносодержащей продукции в неположенном месте',
        4: 'Нецензурная речь'
    }
    c3 = {
        1: 'Пропуск программных моментов без уважительной причины',
        2: 'Оскорбления/конфликты на почве розни'
    }
    match complaintes[user_id].status:
        case 'alert':
            complaintes[user_id].violetion = c1[data]
        case 'soon':
            complaintes[user_id].violetion = c2[data]
        case _:
            complaintes[user_id].violetion = c3[data]
    await callback_query.message.answer("Опишите вашу жалобу подробно.")
    await state.set_state(ComplaintProcess.waiting_for_complaint_text)

@router.message(ComplaintProcess.waiting_for_complaint_text)
async def process_complaint_text(message: Message, state: FSMContext):
    user_id = message.from_user.id
    if user_id in complaintes:
        complaintes[user_id].description = message.text
        await message.answer("При наличии доказательств, прикрепите их в следующем сообщении. Отправьте фото или видео, или /skip чтобы завершить без файлов.")
        await state.set_state(ComplaintProcess.waiting_for_complaint_files)

@router.message(Command("skip"), ComplaintProcess.waiting_for_complaint_files)
async def skip_files(message: Message, state: FSMContext):
    user_id = message.from_user.id
    complaint = complaintes[user_id]
    if not complaint:
        await message.answer("Жалоба не найдена. Начните заново.")
        return

    complaint_id = await add_complaint(complaint)
    await log_complaint_created(complaint.user_id, complaint.adresat, complaint_id)
    user = await get_user(user_id)
    if complaint.status == "alert":
        if user.role != 'Участник':
            await send_complaint_notify(message.bot, complaint, state)
        else:
            await notify_all_reiting_team(message.bot, complaint, state)
    if complaint.status == 'soon':
        await send_complaint_notify_soon(message.bot, complaint, state)
    if complaint.status == 'room_problems':
        await send_complaint_room_problems(message.bot, complaint)
    if complaint.status == 'other':
        await send_complaint_room_problems(message.bot, complaint)
    complaintes.pop(user_id, None)

    await message.answer("Жалоба отправлена. Спасибо.")
    await show_main_menu(message.bot, user_id, state)

async def _finalize_complaint(bot: Bot, user_id: int, state: FSMContext):
    complaint = complaintes[user_id]
    if not complaint:
        await bot.send_message(user_id, "Жалоба не найдена. Начните заново.")
        return

    complaint_id = await add_complaint(complaint)
    await log_complaint_created(complaint.user_id, complaint.adresat, complaint_id)
    await link_files_to_complaint(complaint_id, complaint.files)
    user = await get_user(user_id)
    if complaint.status == "alert":
        if user.role != 'Участник':
            await send_complaint_notify(bot, complaint, state)
        else:
            await notify_all_reiting_team(bot, complaint, state)
    if complaint.status == 'soon':
        await send_complaint_notify_soon(bot, complaint, state)

    complaintes.pop(user_id, None)

    await bot.send_message(user_id, "Жалоба отправлена. Спасибо.")
    await show_main_menu(bot, user_id, state)

async def add_media_to_current_complaint(message: Message, complaint: Complaint):
    saved = 0

    if message.photo:
        if complaint.photo_count < MAX_PHOTOS:
            photo = message.photo[-1]
            file = File(
                id=None,
                tg_id=complaint.user_id,
                tg_file_id=photo.file_id,
                complaint_id=None,
                file_name=f"photo_{complaint.photo_count + 1}.jpg",
                mime_type="image/jpeg",
                file_size=photo.file_size
            )
            file_row_id = await add_file(file)
            file.id = file_row_id
            await log_file_attached(complaint.user_id, file.id, file.tg_file_id)
            complaint.files.append(file.id)
            await log_file_attached(complaint.user_id, file.id, file.tg_file_id)
            complaint.photo_count += 1
            saved += 1

    if message.video:
        if complaint.video_count < MAX_VIDEOS:
            video = message.video
            file = File(
                id=None,
                tg_id=complaint.user_id,
                tg_file_id=video.file_id,
                complaint_id=None,
                file_name=video.file_name or f"video_{complaint.video_count + 1}.mp4",
                mime_type=video.mime_type or "video/mp4",
                file_size=video.file_size
            )
            file_row_id = await add_file(file)
            file.id = file_row_id
            await log_file_attached(complaint.user_id, file.id, file.tg_file_id)
            complaint.files.append(file.id)
            await log_file_attached(complaint.user_id, file.id, file.tg_file_id)
            complaint.video_count += 1
            saved += 1

    if saved == 0:
        return 0, "Прикрепите фото или видео."
    return saved, f"Доказательства приняты: +{saved}."

async def flush_album(user_id: int, media_group_id: str, bot: Bot, chat_id: int, state: FSMContext):
    key = (user_id, media_group_id)
    msgs = _album_buffer.pop(key, [])
    _album_tasks.pop(key, None)

    if not msgs:
        return

    complaint = complaintes.get(user_id)
    if not complaint:
        await bot.send_message(chat_id, "Жалоба не найдена. Начните заново.")
        return

    results = []
    for msg in sorted(msgs, key=lambda m: m.message_id):
        _, text = await add_media_to_current_complaint(msg, complaint)
        results.append(text)

    await bot.send_message(chat_id, "Файлы обработаны:\n" + "\n".join("• " + r for r in results))
    await _finalize_complaint(bot, user_id, state, chat_id)

@router.message(ComplaintProcess.waiting_for_complaint_files)
async def handle_files(message: Message, state: FSMContext):
    user_id = message.from_user.id
    complaint = complaintes.get(user_id)
    if not complaint:
        await message.answer("Жалоба не найдена. Начните заново.")
        return

    if message.media_group_id:
        mgid = str(message.media_group_id)
        key = (user_id, mgid)
        _album_buffer[key].append(message)

        task = _album_tasks.get(key)
        if task and not task.done():
            task.cancel()

        async def delayed():
            try:
                await asyncio.sleep(ALBUM_FLUSH_DELAY)
                await flush_album(user_id, mgid, message.bot, message.chat.id, state)
            except asyncio.CancelledError:
                return
            except Exception as e:
                await message.bot.send_message(message.chat.id, f"Ошибка обработки альбома: {e}")

        _album_tasks[key] = asyncio.create_task(delayed())
        return

    saved, text = await add_media_to_current_complaint(message, complaint)
    if saved == 0:
        await message.answer(text)
        return

    await message.answer("Файлы обработаны. " + text)
    await _finalize_complaint(message.bot, user_id, state)

"""USERS"""

@router.callback_query(MainMenu.users)
async def process_users_callback(callback_query: CallbackQuery, state: FSMContext):
    data = callback_query.data
    user_id = callback_query.from_user.id
    match data:
        case 'back_to_main_menu':
            await show_main_menu(callback_query.bot, user_id, state)
            return
        case 'all_users':
            users = await get_all_users()
            mes = "Список всех пользователей:\n"
            i = 0
            for user in users:
                if i == 39:
                    mes += f"Бейдж: {user.badge_number}, ФИО: {user.fio}, Роль: {user.role}\n"
                    await callback_query.bot.send_message(chat_id=user_id, text=mes, reply_markup=get_users_keyboard())
                    mes = ''
                    i = 0
                else:
                    mes += f"Бейдж: {user.badge_number}, ФИО: {user.fio}, Роль: {user.role}\n"
                    i += 1
            await callback_query.message.answer(mes, reply_markup=get_users_keyboard())
        case 'edit_user_data':
            await callback_query.message.answer('Введите номер бейджа пользователя, данные которого хотите изменить.')
            await state.set_state(UserDataEdit.waiting_for_badge_number)
            await state.update_data(choice=data)
        case 'del_user':
            await callback_query.message.answer('Введите номер бейджа пользователя, которого хотите удалить.')
            await state.set_state(UserDataEdit.waiting_for_badge_number)
            await state.update_data(choice=data)
        case _:
            users = await get_all_users()
            mes = "Список всех пользователей:\n"
            i = 0
            for user in users:
                if i == 39:
                    mes += f"Бейдж: {user.badge_number}, ФИО: {user.fio}, Роль: {user.role}\n"
                    await callback_query.bot.send_message(chat_id=user_id, text=mes, reply_markup=get_users_keyboard())
                    mes = ''
                    i = 0
                else:
                    mes += f"Бейдж: {user.badge_number}, ФИО: {user.fio}, Роль: {user.role}\n"
                    i += 1
            await callback_query.message.answer(mes, reply_markup=get_users_keyboard())

@router.message(UserDataEdit.waiting_for_badge_number)
async def process_user_data_badge(message: Message, state: FSMContext): 
    user_id = message.from_user.id
    data = await state.get_data()
    choice = data.get('choice')
    if not message.text.isdigit():
        await message.answer("Номер должен быть числом, введите еще раз")
        return
    num = int(message.text)
    user = await get_user_by_badge(num)
    if user:
        if user.tg_id == user_id:
            await message.answer("Вы не можете изменить данные самого себя. Введите еще раз бейдж")
            return
        match choice:
            case 'del_user':
                await delete_user(num)
                await message.answer("Пользователь удален.") 
                await show_main_menu(message.bot, user_id, state)
            case 'edit_user_data':
                if user.role == 'Рейтинг':
                    await message.answer("Вы не можете изменить данные этого пользователя.")
                    await state.update_data(choice=choice)
                    return
                edit_users[user_id] = user
                await message.answer(f"Выбран пользователь: {user.fio}. Выберете что будем изменять.", reply_markup=get_edit_badge_keyboard())
                await state.set_state(UserDataEdit.waiting_for_change_choice)
    else:
        await message.answer('Такого человека не существует, введите номер бейджа еще раз')
        await state.update_data(choice=choice)
        return

@router.callback_query(UserDataEdit.waiting_for_change_choice)
async def process_user_data_change(callback_query: CallbackQuery, state: FSMContext):
    user_id = callback_query.from_user.id
    data = callback_query.data
    match data:
        case 'fio':
            await state.update_data(field_to_change='fio')
            await callback_query.message.answer('Введите новое ФИО, все с большой буквы в именительном падеже.')
            await state.set_state(UserDataEdit.waiting_for_new_value)
        case 'team_number':
            await state.update_data(field_to_change='team_number')
            await callback_query.message.answer('Введите новую команду (номер).')
            await state.set_state(UserDataEdit.waiting_for_new_value)
        case 'role':
            await state.update_data(field_to_change='role')
            await callback_query.message.answer('Введите новую роль.')
            await state.set_state(UserDataEdit.waiting_for_new_value)
        case 'badge_number':
            await state.update_data(field_to_change='badge_number')
            await callback_query.message.answer('Введите новый номер бейджа.')
            await state.set_state(UserDataEdit.waiting_for_new_value)
        case 'reiting':
            await state.update_data(field_to_change='reiting')
            await callback_query.message.answer('Введите новое количество очков рейтинга.')
            await state.set_state(UserDataEdit.waiting_for_new_value)
        case 'balance':
            await state.update_data(field_to_change='balance')
            await callback_query.message.answer('Введите новое количество очков баланса.')
            await state.set_state(UserDataEdit.waiting_for_new_value)
        case 'edit_user_back':
            await callback_query.message.answer('Вы вернулись в меню изменения данных пользователя.')
            await show_main_menu(callback_query.bot, user_id, state)
        case _:
            await callback_query.message.answer('Неверный выбор, выберете что будем изменять.')
            await state.set_state(UserDataEdit.waiting_for_change_choice) 

@router.message(UserDataEdit.waiting_for_new_value)
async def process_user_new_value(message: Message, state: FSMContext):
    user_id = message.from_user.id
    user = edit_users.get(user_id)

    if not user:
        await message.answer("Пользователь для редактирования не найден, начните заново.")
        await show_main_menu(message.bot, user_id, state)
        return

    data = await state.get_data()
    field = data.get("field_to_change")
    value = message.text.strip()

    match field:
        case "fio":
            user.fio = value
        case "team_number":
            if not value.isdigit():
                await message.answer("Номер команды должен быть числом, введите еще раз.")
                return
            if int(value) >= 10 or int(value) <= 0:
                await message.answer("Номер команды должен быть от 1 до 9, введите еще раз.")
                return
            user.team_number = int(value)
        case "role":
            if value.lower() not in ["участник", "организатор", "рпг", "администратор", "рейтинг", "медиа"]:
                await message.answer("Неверная роль, введите еще раз.")
                return
            user.role = value
        case "badge_number":
            if 'ком' in value.lower() and len(value) <= 6:
                if value[3:].isdigit():
                    user.badge_number = int(value[3:])
                else:
                    await message.answer("Номер бейджа должен быть числом, введите еще раз.")
                    return
            elif int(value) > 10 or int(value) < 1000:
                user.badge_number = int(value[3:])
            else:
                await message.answer("Неверный формат ввода. Введите данные в виде 'ком123' или число от 10 до 1000.")
                return
        case "reiting":
            if not value.isdigit():
                await message.answer("Рейтинг должен быть числом, введите еще раз.")
                return
            if int(value) < 0 or int(value) > 1000:
                await message.answer("Рейтинг не может быть отрицательным, введите еще раз.")
                return
            user.reiting = int(value)
        case "balance":
            if not value.isdigit():
                await message.answer("Баланс должен быть числом, введите еще раз.")
                return
            if int(value) < 0:
                await message.answer("Баланс не может быть отрицательным, введите еще раз.")
                return
            user.balance = int(value)
        case _:
            await message.answer("Неизвестное поле для изменения, начните заново.")
            await show_main_menu(message.bot, user_id, state)
            return

    await update_user(user)
    active_sessions[user.tg_id] = user

    await message.answer("Данные пользователя обновлены.")
    await show_main_menu(message.bot, user_id, state)

"""REGISTRATION"""

@router.message(Reg.waiting_for_bage_number)
async def process_badge_number(message: Message, state: FSMContext):
    user_id = message.from_user.id
    if message.text.isdigit():
        badge_number = int(message.text)
        existing_user = await get_user_by_badge(badge_number)
        if not existing_user:
            await message.answer("Такого пользователя нет. Введите номер бейджа еще раз.")
            return
    elif 'ком' in message.text.lower() and len(message.text) <= 6:
        if message.text[3:].isdigit():
            badge_number = int(message.text[3:])
            existing_user = await get_user_by_badge(badge_number)
            if not existing_user:
                await message.answer("Такого пользователя нет. Введите номер бейджа еще раз.")
                return
    else:
        await message.answer('Неверный формат ввода. Введите данные в виде')
    registration[user_id].user_id = user_id
    registration[user_id].badge_number = badge_number

    await message.answer("Введите ФИО, все с большой буквы в именительном падеже.")
    await state.set_state(Reg.waiting_for_fio)

@router.message(Reg.waiting_for_fio)
async def process_fio(message: Message, state: FSMContext):
    user_id = message.from_user.id
    registration[user_id].fio = message.text.strip()
    existing_user = await get_user_by_badge(registration[user_id].badge_number)
    if not existing_user:
        await message.answer("Такого пользователя не существует, проверьте введеные данные. Пришлите номер бейджа")
        await state.set_state(Reg.waiting_for_bage_number)
        return
    if message.text.strip() != existing_user.fio:
        await message.answer("Неправильное ФИО, пользователь с такими данными не найден, введите данные еще раз. Номер бейджа")
        await state.set_state(Reg.waiting_for_bage_number)
        return
    if existing_user.tg_id != user_id:
        await update_tg_id(
            registration[user_id].badge_number,
            user_id,
            message.from_user.username,
        )
        try:
            await del_from_active(existing_user.tg_id)
        except csv.Error:
            print("Ну нет так нет")
    try:
        active_sessions[user_id] = USERS[user_id]
    except KeyError:
        print(USERS)
        user = await get_user(user_id)
        active_sessions[user_id] = user
    await message.answer("Регистрация завершена! Спасибо.")
    await add_active(user_id, active_sessions[user_id].role)
    await show_main_menu(message.bot, user_id, state)

"""MAILING"""

async def _send_mailing_payload(bot, tg_id: int, message: Message):
    if message.photo:
        await bot.send_photo(
            tg_id,
            message.photo[-1].file_id,
            caption=((message.caption or "").strip() or None),
        )
        return

    if message.video:
        await bot.send_video(
            tg_id,
            message.video.file_id,
            caption=((message.caption or "").strip() or None),
        )
        return

    if message.voice:
        await bot.send_voice(
            tg_id,
            message.voice.file_id,
            caption=((message.caption or "").strip() or None),
        )
        return

    if message.video_note:
        await bot.send_video_note(tg_id, message.video_note.file_id)
        return

    text = (message.text or "").strip()
    await bot.send_message(tg_id, text)

@router.message(
    Mailing.waiting_for_mailing_text,
    F.content_type.in_({"text", "photo", "video", "voice", "video_note"}))
async def handle_mailing_text(message: Message, state: FSMContext, bot: Bot):
    user_id = message.from_user.id
    user = await get_user(user_id)

    if (
        (message.content_type == "text" and not (message.text or "").strip())
        or (message.content_type != "text" and not (message.caption or "").strip() and not message.video_note)
    ):
        await message.answer("Пришли текст или прикрепи фото/видео/голосовое/кружочек (можно с подписью).")
        return

    match maling[user_id]:
        case "user":
            u = maling_special[user_id]
            if u.tg_id:
                if u.tg_id < 1000:
                    await message.answer("У пользователя подозрительный id, возможно он еще не в системе и письмо не дошло")
                await _send_mailing_payload(message.bot, u.tg_id, message)
                await message.answer("Сообщение отправлено успешно")
            else:
                await message.answer("У этого пользователя пока нет id он не в системе!")

        case "team":
            team = maling_special[user_id]
            users = await get_users_by_team(team.team_number)
            c = 0
            for i in users:
                if i.tg_id:
                    await _send_mailing_payload(message.bot, i.tg_id, message)
                    c += 1
            await message.answer(f"Отправлено {c} сообщений участников комманды из {len(users)}")

        case "all":
            users = await get_all_users()
            c = 0
            for i in users:
                if i.tg_id and i.tg_id > 1000 and i.tg_id in active_sessions:
                    await _send_mailing_payload(message.bot, i.tg_id, message)
                    c += 1
            await message.answer(f"Отправлено {c} сообщений участникам из {len(users)}")

    await show_main_menu(message.bot, user_id, state)

@router.callback_query(Mailing.waiting_adresat)
async def process_maling_choose(message: CallbackQuery, state: FSMContext):
    user_id = message.from_user.id
    data = message.data
    maling[user_id] = data

    match data:
        case "user":
            await message.message.answer("Введите номер бейджа человека")
        case "team":
            await message.message.answer("Введите номер команды")
        case "trek":
            await message.message.answer("Введите номер трека")
        case "all":
            await message.message.answer("Пришли текст или прикрепи фото/видео/голосовое/кружочек (можно с подписью)")
            await state.set_state(Mailing.waiting_for_mailing_text)
            return

    await state.set_state(Mailing.waiting_info)

@router.message(Mailing.waiting_info)
async def process_maling_adresat(message: Message, state: FSMContext):
    user_id = message.from_user.id
    text = message.text or ""

    match maling[user_id]:
        case "user":
            if text.isdigit():
                user = await get_user_by_badge(int(text))
                if not user:
                    await message.answer("Такого пользователя не существуют попробуйте еще раз")
                    return
                await message.answer("Пришли текст или прикрепи фото/видео/голосовое/кружочек (можно с подписью)")
                maling_special[user_id] = user
            else:
                await message.answer("Номер бейджа должен быть числом, пришлите еще раз")
                return

        case "team":
            if text.isdigit():
                team = await get_team(int(text))
                if not team:
                    await message.answer("Такой команды не существует попробуйте еще раз")
                    return
                await message.answer("Пришли текст или прикрепи фото/видео/голосовое/кружочек (можно с подписью)")
                maling_special[user_id] = team
            else:
                await message.answer("Номер бейджа должен быть числом, пришлите еще раз")
                return

        case "trek":
            if text.isdigit():
                await message.answer("Не, я пока балдеЮ")
                await show_main_menu(message.bot, user_id, state)
                return
            else:
                await message.answer("Номер бейджа должен быть числом, пришлите еще раз")
                return

    await state.set_state(Mailing.waiting_for_mailing_text)

@router.callback_query(YesNoChoice.waiting_for_message_answer)
async def process_message_answering(message: CallbackQuery, state: FSMContext):
    user_id = message.from_user.id
    m = messagess[user_id]
    data = message.data
    match data:
        case 'seen':
            await update_status(m.id, 'answered')
            messages = await get_new_messages()
            if messages:
                m = messages[0]
                messagess[user_id] = m
                us = await get_user(m.user_id)
                await message.answer(f'Сообщение от {us.fio} {us.username}\n\n' + m.text, reply_markup=get_message_keyboard())
            await state.set_state(YesNoChoice.waiting_for_message_answer)
        case 'skip':
            await update_status(m.id, 'skip')
            messages = await get_new_messages()
            if messages:
                m = messages[0]
                messagess[user_id] = m
                us = await get_user(m.user_id)
                await message.answer(f'Сообщение от {us.fio} {us.username}\n\n' + m.text, reply_markup=get_message_keyboard())
            await state.set_state(YesNoChoice.waiting_for_message_answer)

"""RATING"""

@router.message(Rating.waiting_for_badge_number)
async def handle_rating_badge_number(message: Message, state: FSMContext):
    user_id = message.from_user.id
    badge_number = message.text
    if not badge_number.isdigit():
        await message.answer('Введите корректный бейдж')
        return
    user = await get_user_by_badge(int(badge_number))
    if not user:
        await message.answer('Такого пользователя не существует попробуйте еще раз')
        return
    await message.answer('Выберете действие', reply_markup=get_rating_choice_keyboard())
    await state.set_state(Rating.waiting_for_choice)
    await state.update_data(badge_number=badge_number)

@router.callback_query(Rating.waiting_for_choice)
async def handle_rating_choice(callback: CallbackQuery, state: FSMContext):
    user_id = callback.from_user.id
    data = callback.data
    await state.update_data(choice=data)
    match data:
        case 'add':
            await callback.message.answer('Введите сумму начисления')
            await state.set_state(Rating.waiting_for_amount)
        case 'subtract':
            await callback.message.answer('Введите сумму штрафа')
            await state.set_state(Rating.waiting_for_amount)

@router.message(Rating.waiting_for_amount)
async def handle_rating_amount(message: Message, state: FSMContext):
    user_id = message.from_user.id
    amount = message.text
    data = await state.get_data()
    badge_number = data.get('badge_number')
    if not amount.isdigit():
        await message.answer('Введите корректную сумму')
        return
    match data.get('choice'):
        case 'add':
            await add_rating(badge_number, int(amount))
        case 'subtract':
            await subtract_rating(badge_number, int(amount))
    await notify_user_reiting(message.bot, data.get('choice'), badge_number, int(amount))
    await message.answer('Действие выполнено')
    await show_main_menu(message.bot, user_id, state)
    await state.set_state(MainMenu.main_menu_rating_team)

@router.message(Rating.waiting_for_badge_number_bonus)
async def give_bonus_badge_number(message: Message, state: FSMContext):
    user_id = message.from_user.id
    amount = message.text
    if not amount.isdigit():
        await message.answer('Введите корректный бейдж')
        return
    badge_number = int(amount)
    await message.answer('Введите сумму бонуса')
    await state.update_data(choice=badge_number)
    await state.set_state(Rating.waiting_for_amount_bonus)

@router.message(Rating.waiting_for_amount_bonus)
async def give_bonus_amount(message: Message, state: FSMContext):
    user_id = message.from_user.id
    amount_text = message.text

    if not amount_text.isdigit():
        await message.answer("Введите корректную сумму бонуса")
        return

    amount = int(amount_text)
    if amount <= 0:
        await message.answer("Сумма бонуса должна быть больше 0")
        return

    data = await state.get_data()
    badge_number = data.get("choice")

    if not badge_number:
        await message.answer("Ошибка: не найден номер бейджа")
        await state.clear()
        return

    await add_rating(badge_number, amount)

    await message.answer(
        f"Бонус {amount} успешно начислен участнику с бейджем {badge_number}"
    )

    await show_main_menu(message.bot, user_id, state)
    
"""PROMOKODS"""

@router.message(PromoCreate.waiting_for_phrase)
async def process_promo_phrase(message: Message, state: FSMContext):
    user_id = message.from_user.id
    phrase = message.text
    user = await get_user(user_id)
    if user:
        promos[user_id] = Promokod(phrase=phrase, badge_number=user.badge_number)
        await message.answer('Введите колличество промокодов')
        await state.set_state(PromoCreate.waiting_for_amount)
    else:
        await show_main_menu(message.bot, user_id, state)

@router.message(PromoCreate.waiting_for_amount)
async def process_promo_amount(message: Message, state: FSMContext):
    user_id = message.from_user.id
    text = message.text
    if not text.isdigit():
        await message.answer('Кол-во должно быть числом')
        await state.set_state(PromoCreate.waiting_for_amount)
        return
    
    promos[user_id].amount = int(text)
    await message.answer('Введите бонус')
    await state.set_state(PromoCreate.waiting_for_bonus)

@router.message(PromoCreate.waiting_for_bonus)
async def process_promo_bonus(message: Message, state: FSMContext):
    user_id = message.from_user.id
    text = message.text
    if not text.isdigit():
        await message.answer('Бонус должен быть числом')
        await state.set_state(PromoCreate.waiting_for_amount)
        return
    
    promos[user_id].bonus = int(text)
    i = await add_promokod(promos[user_id])
    await message.answer(f'Промокод номер {i} {promos[user_id].bonus} добавлен')
    del promos[user_id]
    await show_main_menu(message.bot, user_id, state)

@router.message(PromoCreate.waiting_for_phrase_user)
async def process_promo_phrase_user(message: Message, state: FSMContext):
    user_id = message.from_user.id
    phrase = message.text
    user = await get_user(user_id)
    promokod = await get_promo_by_pharse(phrase)
    if promokod:
        await message.answer(f'Промокод применен, на баланс начисленноо {promokod.bonus} единиц')
        await update_promokod(promokod.id, promokod.bonus, promokod.amount - 1)
    else:
        await message.answer('Такого промокода не существует либо он закончился')
    await callback_query.message.answer("Магазин.", reply_markup=get_student_shop_keyboard())
    await state.set_state(Shop.wait_for_choice_action)

"""BONUS"""

@router.callback_query(Bonus.waiting_adresat)
async def process_bonus_adresat(callback_query: CallbackQuery, state: FSMContext):
    user_id = callback_query.from_user.id
    data = callback_query.data
    await state.update_data(choice=data)
    match data:
        case 'user':
            await callback_query.message.answer('Введите номер бейджа человека')
        case 'team':
            await callback_query.message.answer('Введите номер команды')
        case 'all':
            await callback_query.message.answer('Жду текст сообщения')
            await state.set_state(Bonus.waiting_amount)
            return
    await state.set_state(Bonus.waiting_number)

@router.message(Bonus.waiting_number)
async def process_bonus_adresat(message: Message, state: FSMContext):
    user_id = message.from_user.id
    text = message.text
    if not text.isdigit():
        await message.answer('Номер должен быть числом введите еще раз')
        return
    await state.update_data(number=int(text))
    await message.answer('Введите размер бонуса')
    await state.set_state(Bonus.waiting_amount)

@router.message(Bonus.waiting_amount)
async def process_bonus_amount(message: Message, state: FSMContext):
    user_id = message.from_user.id
    text = message.text

    if not text.isdigit():
        await message.answer("Кол-во должно быть числом, введите еще раз")
        return

    amount = int(text)
    if amount <= 0:
        await message.answer("Сумма должна быть больше 0")
        return

    data = await state.get_data()
    choice = data.get("choice")     # "user" | "team" | "all"
    number = data.get("number")     # badge_number или team_number (для user/team)

    match choice:
        case "user":
            if not number:
                await message.answer("Ошибка: не найден номер бейджа")
                await state.clear()
                return
            await add_bonus(number, amount)
            await message.answer(f"Начислил бонус {amount} участнику с бейджем {number}")

        case "team":
            if not number:
                await message.answer("Ошибка: не найден номер команды")
                await state.clear()
                return
            users = await get_users_by_team(number)
            if not users:
                await message.answer("В этой команде нет пользователей")
                await state.clear()
                return
            for u in users:
                if u.badge_number is not None:
                    await add_bonus(u.badge_number, amount)
            await message.answer(f"Начислил бонус {amount} всем участникам команды {number}")

        case "all":
            users = await get_all_users()
            if not users:
                await message.answer("Пользователей нет")
                await state.clear()
                return
            for u in users:
                if u.badge_number is not None:
                    await add_rating(u.badge_number, amount)
            await message.answer(f"Начислил бонус {amount} всем участникам")

        case _:
            await message.answer("Ошибка: неизвестный получатель")
            await state.clear()
            return
        
    await show_main_menu(message.bot, user_id, state)
    await state.clear()

"""PRODUCTS"""
@router.callback_query(Products.wait_choice_action)
async def process_product_choice(callback_query: CallbackQuery, state: FSMContext):
    user_id = callback_query.from_user.id
    await state.update_data(choice=callback_query.data)
    await callback_query.message.answer('Введите название продукта')
    await state.set_state(Products.wait_for_product_name)

@router.message(Products.wait_for_product_name)
async def process_product_name(message: Message, state: FSMContext):
    user_id = message.from_user.id
    data = await state.get_data()
    prod = await get_product_by_name(message.text)

    match data.get('choice'):
        case 'add':
            if prod:
                await message.answer('Такой продукт уже существует введите название еще раз')
                return
            await state.update_data(name=message.text)
            await message.answer('Введите стоимость товара')
            await state.set_state(Products.wait_for_product_cost)
        case 'update':
            if not prod:
                await message.answer('Такого продукта не существует введите название еще раз')
                return
            await message.answer('Выберете изменение поля', reply_markup=get_product_edit_keyboard())
            await state.set_state(Products.wait_choice_edit)
        case 'del':
            if not prod:
                await message.answer('Такого продукта не существует введите название еще раз')
                return
            await delete_product(prod.id)
            await message.answer('Товар успешно удален')
            await show_main_menu(message.bot, user_id, state)

@router.message(Products.wait_for_product_cost)
async def process_product_cost(message: Message, state: FSMContext):
    user_id = message.from_user.id
    text = message.text
    if not text.isdigit():
        await message.answer('Стоимость должна быть числом введите еще раз')
        return
    await state.update_data(cost=int(text))
    await message.answer('Введите кол-во товаров')
    await state.set_state(Products.wait_for_product_amount)

@router.message(Products.wait_for_product_amount)
async def process_product_amount(message: Message, state: FSMContext):
    user_id = message.from_user.id
    data = await state.get_data()
    text = message.text
    if not text.isdigit():
        await message.answer('Кол-во должно быть числом введите еще раз')
        return
    
    name = data.get('name')
    cost = data.get('cost')
    amount = int(text)
    i = await add_product(Product(name=name, cost=cost, amount=amount))
    await message.answer(f'Продукт номер {i}: {name} успешно добавлен')
    await message.bot.send_message(user_id, 'Выберете дейстивие', reply_markup=get_edit_product_choice())
    await state.set_state(Products.wait_choice_action)

@router.callback_query(Products.wait_choice_edit)
async def process_product_choice_edit(callback_query: CallbackQuery, state: FSMContext):
    user_id = callback_query.from_user.id
    data = callback_query.data  # формат: "name:ID" | "cost:ID" | "amount:ID"

    try:
        action, product_id = data.split(":")
        product_id = int(product_id)
    except ValueError:
        await callback_query.answer("Ошибка выбора", show_alert=True)
        return

    product = await get_product(product_id)
    if not product:
        await callback_query.answer("Товар не найден", show_alert=True)
        return

    await state.update_data(
        product_id=product_id,
        edit_action=action,
    )

    match action:
        case "name":
            await callback_query.message.answer("Введите новое название товара")
        case "cost":
            await callback_query.message.answer("Введите новую цену товара")
        case "amount":
            await callback_query.message.answer("Введите новое количество товара")
        case _:
            await callback_query.answer("Неизвестное действие", show_alert=True)
            return

    await callback_query.answer()
    await state.set_state(Products.wait_for_edit_value)

@router.message(Products.wait_for_edit_value)
async def process_product_edit_value(message: Message, state: FSMContext):
    data = await state.get_data()
    product_id = data.get("product_id")
    action = data.get("edit_action")

    product = await get_product(product_id)
    if not product:
        await message.answer("Товар не найден")
        await state.clear()
        return

    value = message.text.strip()

    match action:
        case "name":
            if not value:
                await message.answer("Название не может быть пустым")
                return
            product.name = value

        case "cost":
            if not value.isdigit():
                await message.answer("Цена должна быть числом")
                return
            product.cost = int(value)

        case "amount":
            if not value.isdigit():
                await message.answer("Количество должно быть числом")
                return
            product.amount = int(value)

    await update_product(product)

    await message.answer("Товар успешно обновлён")
    await state.clear()
    await show_main_menu(message.bot, message.from_user.id, state)

"""SHOP"""
@router.callback_query(Shop.wait_for_choice_action)
async def process_shop_choice(callback_query: CallbackQuery, state: FSMContext):
    user_id = callback_query.from_user.id
    data = callback_query.data
    user = await get_user(user_id)
    match data:
        case 'products':
            products = await get_products_shop()
            if not products:
                await callback_query.message.answer('Пока магазин пуст, возвращайтесь позже')
                await show_main_menu(callback_query.bot, user_id, state)
            if user:
                await callback_query.message.answer(f'Ваш баланс: {user.balance}')
            last = products[-1]
            products.remove(last)
            if products:
                for i in products:
                    await callback_query.bot.send_message(user_id, i)
            await callback_query.bot.send_message(user_id, last, reply_markup=get_buy_choice())
            await state.set_state(Shop.wait_for_buy_choice)
        case 'my_buy':
            buys = await get_my_purchases(user.badge_number)
            if not buys:
                await callback_query.message.answer('Пока что не было покупок')
                await callback_query.message.answer("Магазин.", reply_markup=get_student_shop_keyboard())
                await state.set_state(Shop.wait_for_choice_action)
                return
            await callback_query.bot.send_message(user_id, buys)
            await callback_query.message.answer("Магазин.", reply_markup=get_student_shop_keyboard())
            await state.set_state(Shop.wait_for_choice_action)
        case 'tasks':
            await callback_query.message.answer('Скоро тут появятся задания, их можно будет выполнять')
            await callback_query.message.answer("Магазин.", reply_markup=get_student_shop_keyboard())
            await state.set_state(Shop.wait_for_choice_action)
        case 'give_promo':
            await callback_query.message.answer('Введите промокод')
            await state.set_state(PromoCreate.waiting_for_phrase_user)
        case 'back_to_main_menu':
            await show_main_menu(callback_query.bot, user_id, state)

@router.callback_query(Shop.wait_for_buy_choice)
async def process_buy_choice(callback_query: CallbackQuery, state: FSMContext):
    user_id = callback_query.from_user.id
    data = callback_query.data
    if data == 'back_to_main_menu':
        await show_main_menu(callback_query.bot, user_id, state)
        return
    await state.update_data(adresat=data)
    if data == 'for_me':
        await callback_query.message.answer('Введите номер товара, который хотите купить')
        await state.set_state(Shop.wait_for_product_number)
    else:
        await callback_query.message.answer('Введите номер бейджа кому подарок или ФИО')
        await state.set_state(Shop.wait_for_user_data)

@router.message(Shop.wait_for_user_data)
async def process_buy_user_data(message: Message, state: FSMContext):
    user_id = message.from_user.id
    text = message.text.strip()

    data = await state.get_data()
    adresat = data.get("adresat")  # ожидается не 'for_me'

    # пытаемся определить адресата: сначала бейдж, потом ФИО
    badge_number = None
    target_user = None

    if text.isdigit():
        badge_number = int(text)
        target_user = await get_user_by_badge(badge_number)
    else:
        target_user = await get_user_by_fio(text)

    if not target_user or not target_user.badge_number:
        await message.answer(
            "Пользователь не найден. Введите номер бейджа или ФИО ещё раз"
        )
        return

    await state.update_data(
        target_badge=target_user.badge_number
    )

    await message.answer("Введите номер товара, который хотите подарить")
    await state.set_state(Shop.wait_for_product_number)

@router.message(Shop.wait_for_product_number)
async def process_product_number(message: Message, state: FSMContext):
    user_id = message.from_user.id
    text = message.text
    data = await state.get_data()
    adr = None
    match data.get('choice'):
        case 'gift':
            adr = await get_user_by_badge(data.get('target_badge'))
    if not text.isdigit():
        await message.answer('Номер должен быть числом введите еще раз')
        return
    
    user = await get_user(user_id)
    balance = await get_user_balance(user.badge_number)
    product = await get_product(int(text))
    if not product:
        await message.answer('Продукта с таким номер не существует, введите еще раз')
        return
    if user:
        if balance >= product.cost:
            if adr:
                await message.answer(f'Вы хотите приобрести {product.name} за {product.cost} для {adr.fio}?', reply_markup=get_buy_keyboard())
                await state.update_data(adres=adr.badge_number)
            else:
                await message.answer(f'Вы хотите приобрести {product.name} за {product.cost}?', reply_markup=get_buy_keyboard())
                await state.update_data(adres=user.badge_number)
            await state.update_data(product=int(text))
            await state.set_state(Shop.wait_complite_buy)
        else:
            await message.answer('К сожалению у вас недостаточно средств')
            products = await get_products_shop()
            if user:
                await message.answer(f'Ваш баланс: {balance}')
            last = products[-1]
            products.remove(last)
            if products:
                for i in products:
                    await message.bot.send_message(user_id, i)
            await message.bot.send_message(user_id, last, reply_markup=get_buy_choice())
            await state.update_data(balance=balance)
            await state.set_state(Shop.wait_for_buy_choice)
    else:
        await message.answer('Пользователь не найден, обратитесь к организаторам')
        await show_main_menu(message.bot, user_id, state)

@router.callback_query(Shop.wait_complite_buy)
async def process_complite_buy(callback_query: CallbackQuery, state: FSMContext):
    user_id = callback_query.from_user.id
    data = callback_query.data
    dat = await state.get_data()
    product = await get_product(dat.get('product'))
    user = await get_user(user_id)
    adr = await get_user_by_badge(dat.get('adres')) if dat.get('adres') != user.badge_number else None
    match data:
        case 'buy':
            await buy_product(product.cost, user.badge_number)
            await product_sold(product.id, user.badge_number)
            if adr:
                await notify_rpg_buy(callback_query.bot, adr, product)
            else:
                await notify_rpg_buy(callback_query.bot, user, product)
            await callback_query.message.answer('Товар приобретен!')
        case _:
            await callback_query.message.answer('Покупка отменена')
        
    await show_main_menu(callback_query.bot, user_id, state)    

"""ZAGS"""
@router.callback_query(ZAGS.waiting_for_choice)
async def process_zags_choice(callback_query: CallbackQuery, state: FSMContext):
    user_id = callback_query.from_user.id
    data = callback_query.data
    match data:
        case 'married':
            await callback_query.message.answer('Введите номер бейджа кому предложение')
            await state.set_state(ZAGS.waiting_for_badge)
        case _:
            await show_main_menu(callback_query.bot, user_id, state)

@router.callback_query(ZAGS.waiting_for_badge)
async def process_zags_badge(message: Message, state: FSMContext):
    user_id = callback_query.from_user.id
    text = message.text
    if not text.isdigit():
        await message.answer('Номер должен быть числом введите еще раз')
        return 
    user = await get_user(user_id)
    adr = await get_user_by_badge(int(text))
    if not adr:
        await message.answer('Такого человека не существует')
        return 
    if adr.gender == user.gender:
        await message.answer('Браки могут составляться только между мужчиной и женщиной')
        return 
    await message.answer('Чью фамилию выбираете', reply_markup=get_married_second_name())
    await state.update_data(int(text))
    await state.set_state(ZAGS.waiting_for_fio)

@router.callback_query(ZAGS.waiting_for_fio)
async def process_(callback_query: CallbackQuery, state: FSMContext):
    user_id = callback_query.from_user.id
    data = callback_query.data
    dat = await state.get_data()
    badge_number = dat.get('badge_number')
    adr = await get_user_by_badge(badge_number)
    

"""UPLOAD/EXPORT CSV FILES"""
UPLOAD_RATING_PARTICIPANTS = "upload_rating_participants"
UPLOAD_RATING_TEAMS = "upload_rating_teams"
UPLOAD_PARTICIPANTS = "upload_participants"

EXPORT_RATING_PARTICIPANTS = "export_rating_participants"
EXPORT_RATING_TEAMS = "export_rating_teams"
EXPORT_PARTICIPANTS = "export_participants"
EXPORT_LOGS = "export_logs"

def _parse_int(v, default=0):
    try:
        if v is None:
            return default
        s = str(v).strip()
        if s == "":
            return default
        return int(float(s.replace(" ", "").replace(",", ".")))
    except Exception:
        return default

def _parse_text(v):
    if v is None:
        return ""
    return str(v).strip()

def _read_csv_rows(text: str, delimiter: str) -> list[list[str]]:
    reader = csv.reader(io.StringIO(text), delimiter=delimiter)
    return [r for r in reader if any(str(x).strip() for x in r)]

def _rows_from_csv_bytes(data: bytes) -> list[dict]:
    text = data.decode("utf-8-sig", errors="replace")
    all_rows = _read_csv_rows(text, ";")
    if not all_rows:
        return []
    if all(len(r) <= 1 for r in all_rows) and "," in text:
        all_rows = _read_csv_rows(text, ",")

    header = [c.strip() for c in all_rows[0]]
    expected = ["badge_number", "full_name", "team_id", "daily_base", "penalties_sum", "bonuses_sum", "total_points", "updated_at"]

    has_header = False
    if len(header) >= 7:
        lower = [h.lower() for h in header]
        if all(x in lower for x in expected[:7]):
            has_header = True

    start_idx = 1 if has_header else 0
    rows = []

    for r in all_rows[start_idx:]:
        r = list(r) + [""] * (8 - len(r))
        badge_number = _parse_int(r[0], default=-1)
        if badge_number <= 0:
            continue

        updated_at = _parse_text(r[7]) or now_iso()

        rows.append(
            {
                "badge_number": badge_number,
                "full_name": _parse_text(r[1]),
                "team_id": _parse_int(r[2], default=None),
                "daily_base": _parse_int(r[3], default=100),
                "penalties_sum": _parse_int(r[4], default=0),
                "bonuses_sum": _parse_int(r[5], default=0),
                "total_points": _parse_int(r[6], default=0),
                "updated_at": updated_at,
            }
        )

    return rows

def _rows_from_rating_teams_csv_bytes(data: bytes) -> list[dict]:
    text = data.decode("utf-8-sig", errors="replace")
    all_rows = _read_csv_rows(text, ";")
    if not all_rows:
        return []
    if all(len(r) <= 1 for r in all_rows) and "," in text:
        all_rows = _read_csv_rows(text, ",")

    header = [c.strip().lower() for c in all_rows[0]]
    expected = ["team_number", "team_name", "team_total_points", "updated_at"]
    has_header = all(x in header for x in expected[:2])

    start_idx = 1 if has_header else 0
    rows = []

    for r in all_rows[start_idx:]:
        r = list(r) + [""] * (4 - len(r))
        if has_header:
            header_map = {name: idx for idx, name in enumerate(header)}
            team_number = _parse_int(r[header_map.get("team_number", 0)], default=-1)
            team_name = _parse_text(r[header_map.get("team_name", 1)])
            team_total_points = _parse_int(r[header_map.get("team_total_points", 2)], default=0)
            updated_at = _parse_text(r[header_map.get("updated_at", 3)]) or now_iso()
        else:
            team_number = _parse_int(r[0], default=-1)
            team_name = _parse_text(r[1])
            team_total_points = _parse_int(r[2], default=0)
            updated_at = _parse_text(r[3]) or now_iso()

        if team_number <= 0:
            continue

        rows.append(
            {
                "team_number": team_number,
                "team_name": team_name,
                "team_total_points": team_total_points,
                "updated_at": updated_at,
            }
        )

    return rows

def _rows_from_participants_csv_bytes(data: bytes) -> list[dict]:
    text = data.decode("utf-8-sig", errors="replace")
    all_rows = _read_csv_rows(text, ";")
    if not all_rows:
        return []
    if all(len(r) <= 1 for r in all_rows) and "," in text:
        all_rows = _read_csv_rows(text, ",")

    header = [c.strip().lower() for c in all_rows[0]]
    has_header = any(x in header for x in ("badge", "badge_number")) and "fio" in header
    header_map = {name: idx for idx, name in enumerate(header)} if has_header else {}

    start_idx = 1 if has_header else 0
    rows = []

    for r in all_rows[start_idx:]:
        r = list(r) + [""] * (4 - len(r))

        if has_header:
            badge_number = _parse_int(
                r[header_map.get("badge", header_map.get("badge_number", 0))],
                default=-1,
            )
            fio = _parse_text(r[header_map.get("fio", 1)])
            role = _parse_text(r[header_map.get("role", 2)]) or "Участник"
            gender_raw = _parse_text(r[header_map.get("gender", 3)])
        else:
            badge_number = _parse_int(r[0], default=-1)
            fio = _parse_text(r[1])
            role = _parse_text(r[2]) or "Участник"
            gender_raw = _parse_text(r[3])

        if badge_number < 0:
            continue

        gender = gender_raw if gender_raw in ("М", "Ж") else None

        rows.append(
            {
                "user_id": badge_number,
                "fio": fio,
                "team_number": None,
                "role": role,
                "badge_number": badge_number,
                "gender": gender,
                "reiting": 0,
                "balance": 0,
                "date_registered": now_iso(),
            }
        )

    return rows

@router.message(Command("upload_file"))
async def upload_reiting_cmd(message: Message, state: FSMContext):
    user = await get_user(message.from_user.id)
    role = user.role
    if not role == 'Рейтинг':
        await message.answer("Доступно только для роли «Рейтинг».")
        return
    await state.set_state(RatingCSV.waiting_for_upload_choice)
    await message.answer("Выбери тип файла для загрузки.", reply_markup=get_upload_csv_keyboard())

@router.callback_query(RatingCSV.waiting_for_upload_choice, lambda c: c.data in {UPLOAD_RATING_PARTICIPANTS, UPLOAD_RATING_TEAMS, UPLOAD_PARTICIPANTS})
async def select_upload_csv_type(callback_query: CallbackQuery, state: FSMContext):
    user = await get_user(callback_query.from_user.id)
    role = user.role
    if not role == 'Рейтинг':
        await state.clear()
        await callback_query.message.answer("Доступно только для роли «Рейтинг».")
        return

    await callback_query.answer()
    await state.update_data(upload_type=callback_query.data)
    await state.set_state(RatingCSV.waiting_for_csv)

    if callback_query.data == UPLOAD_RATING_PARTICIPANTS:
        prompt = "Пришли .csv файл с рейтингом участников (разделитель – «;»)."
    elif callback_query.data == UPLOAD_RATING_TEAMS:
        prompt = "Пришли .csv файл с рейтингом команд (team_number; team_name; team_total_points; updated_at)."
    else:
        prompt = "Пришли .csv файл с участниками (badge; fio; role)."

    await callback_query.message.answer(prompt)

@router.message(RatingCSV.waiting_for_csv, F.document)
async def upload_reiting_file(message: Message, state: FSMContext):
    user = await get_user(message.from_user.id)
    role = user.role
    if not role == 'Рейтинг':
        await state.clear()
        await message.answer("Доступно только для роли «Рейтинг».")
        return

    doc = message.document
    name = (doc.file_name or "").lower()
    if not name.endswith(".csv"):
        await message.answer("Нужен файл .csv.")
        return

    file = await message.bot.get_file(doc.file_id)
    data = await message.bot.download_file(file.file_path)
    content = data.read()

    state_data = await state.get_data()
    upload_type = state_data.get("upload_type")
    if not upload_type:
        await message.answer("Сначала выбери тип загрузки командой /upload_reiting.")
        return

    if upload_type == UPLOAD_RATING_PARTICIPANTS:
        rows = _rows_from_csv_bytes(content)
        if not rows:
            await message.answer("Не нашёл валидных строк. Проверь формат файла.")
            return
        n = await upsert_rating_rows(rows)
        await recalc_team_totals()
    elif upload_type == UPLOAD_RATING_TEAMS:
        rows = _rows_from_rating_teams_csv_bytes(content)
        if not rows:
            await message.answer("Не нашёл валидных строк. Проверь формат файла.")
            return
        n = await upsert_rating_team_rows(rows)
    else:
        rows = _rows_from_participants_csv_bytes(content)
        if not rows:
            await message.answer("Не нашёл валидных строк. Проверь формат файла.")
            return
        print(rows)
        for row in rows:
            await add_user(
                User(
                    tg_id=row["user_id"],
                    username=None,
                    fio=row["fio"],
                    team_number=row["team_number"],
                    role=row["role"],
                    badge_number=row["badge_number"],
                    reiting=row["reiting"],
                    balance=row["balance"],
                    date_registered=row["date_registered"],
                )
            )

    await state.clear()
    await message.answer(f"Загружено строк: {len(rows)}.")
    await show_main_menu(message.bot, user.tg_id, state)

@router.message(RatingCSV.waiting_for_csv)
async def upload_reiting_wrong(message: Message):
    await message.answer("Пришли .csv файлом (документом).")

@router.message(RatingCSV.waiting_for_upload_choice)
async def upload_reiting_need_choice(message: Message):
    await message.answer("Сначала выбери тип файла для загрузки.", reply_markup=get_upload_csv_keyboard())

@router.message(Command("get_file"))
async def export_reiting(message: Message, state: FSMContext):
    user = await get_user(message.from_user.id)
    role = user.role
    if not role == 'Рейтинг':
        await message.answer("Доступно только для роли «Рейтинг».")
        return
    await state.set_state(RatingCSV.waiting_for_export_choice)
    await message.answer("Выбери тип выгрузки.", reply_markup=get_export_csv_keyboard())

@router.callback_query(RatingCSV.waiting_for_export_choice, lambda c: c.data in {EXPORT_RATING_PARTICIPANTS, EXPORT_RATING_TEAMS, EXPORT_PARTICIPANTS, EXPORT_LOGS})
async def export_reiting_choice(callback_query: CallbackQuery, state: FSMContext):
    user = await get_user(callback_query.from_user.id)
    role = user.role
    if not role == 'Рейтинг':
        await state.clear()
        await callback_query.message.answer("Доступно только для роли «Рейтинг».")
        return

    await callback_query.answer()
    await state.clear()
    export_choice = callback_query.data

    if export_choice == EXPORT_RATING_PARTICIPANTS:
        await _export_rating_participants(callback_query.message)
    elif export_choice == EXPORT_RATING_TEAMS:
        await _export_rating_teams(callback_query.message)
    elif export_choice == EXPORT_PARTICIPANTS:
        await _export_participants(callback_query.message)
    else:
        await _export_logs(callback_query.message)

@router.message(RatingCSV.waiting_for_export_choice)
async def export_reiting_need_choice(message: Message):
    await message.answer("Сначала выбери тип выгрузки.", reply_markup=get_export_csv_keyboard())

async def export_sells_xlsx(bot: Bot, chat_id: int):
    async with aiosqlite.connect(DB_PATH) as db:
        db.row_factory = aiosqlite.Row

        # на случай опечатки poduct_id vs product_id
        cur = await db.execute("PRAGMA table_info(sells);")
        cols = {r["name"] for r in await cur.fetchall()}
        product_col = "product_id" if "product_id" in cols else ("poduct_id" if "poduct_id" in cols else None)
        if product_col is None:
            raise ValueError("В таблице sells нет колонки product_id/poduct_id")

        cursor = await db.execute(
            f"""
            SELECT
                u.fio AS fio,
                s.badge_number AS badge_number,
                s.{product_col} AS product_id,
                p.name AS product_name,
                s.date_created AS date_created
            FROM sells s
            LEFT JOIN users u ON u.badge_number = s.badge_number
            LEFT JOIN products p ON p.id = s.{product_col}
            ORDER BY s.date_created DESC, s.id DESC;
            """
        )
        rows = await cursor.fetchall()

    wb = Workbook()
    ws = wb.active
    ws.title = "sells"

    headers = ["fio", "badge_number", "product_id", "product.name", "date_created"]
    ws.append(headers)

    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    for r in rows:
        ws.append([
            r["fio"] or "",
            r["badge_number"],
            r["product_id"],
            r["product_name"] or "",
            r["date_created"],
        ])

    widths = [28, 14, 12, 28, 20]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)

    await bot.send_document(
        chat_id=chat_id,
        document=BufferedInputFile(bio.getvalue(), filename="sells.xlsx"),
        caption="Выгрузка продаж (sells).",
    )

async def export_remaining_products_xlsx(bot: Bot, chat_id: int):
    async with aiosqlite.connect(DB_PATH) as db:
        cursor = await db.execute(
            """
            SELECT id, name, cost, amount
            FROM products
            WHERE amount > 0
            ORDER BY id;
            """
        )
        rows = await cursor.fetchall()

    wb = Workbook()
    ws = wb.active
    ws.title = "products"

    headers = ["id", "name", "cost", "amount"]
    ws.append(headers)

    for col in range(1, len(headers) + 1):
        ws.cell(row=1, column=col).font = Font(bold=True)
        ws.column_dimensions[get_column_letter(col)].width = [8, 30, 12, 12][col - 1]

    for row in rows:
        ws.append(row)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    await bot.send_document(
        chat_id=chat_id,
        document=BufferedInputFile(
            buffer.getvalue(),
            filename="remaining_products.xlsx",
        ),
        caption="Оставшиеся товары в магазине",
    )

async def _export_rating_participants(message: Message):
    async with aiosqlite.connect(DB_PATH) as db:
        db.row_factory = aiosqlite.Row
        await db.execute("PRAGMA foreign_keys=ON;")
        cur = await db.execute(
            """
            SELECT badge_number, full_name, team_id, daily_base,
                   penalties_sum, bonuses_sum, total_points, updated_at
            FROM ratings
            ORDER BY team_id, total_points DESC, full_name
            """
        )
        rows = await cur.fetchall()

    output = io.StringIO()
    writer = csv.writer(output, delimiter=";", lineterminator="\n")
    writer.writerow(["badge_number", "full_name", "team_id", "daily_base", "penalties_sum", "bonuses_sum", "total_points", "updated_at"])
    for r in rows:
        writer.writerow(
            [
                r["badge_number"],
                r["full_name"],
                r["team_id"] if r["team_id"] is not None else "",
                r["daily_base"],
                r["penalties_sum"],
                r["bonuses_sum"],
                r["total_points"],
                r["updated_at"] or "",
            ]
        )

    filename = f"reiting_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
    path = f"/tmp/{filename}"
    with open(path, "wb") as f:
        f.write(output.getvalue().encode("utf-8-sig"))

    await message.answer_document(FSInputFile(path, filename=filename))

async def _export_rating_teams(message: Message):
    async with aiosqlite.connect(DB_PATH) as db:
        db.row_factory = aiosqlite.Row
        await db.execute("PRAGMA foreign_keys=ON;")
        cur = await db.execute(
            """
            SELECT team_number, team_name, team_total_points, updated_at
            FROM ratingteams
            ORDER BY team_total_points DESC, team_number
            """
        )
        rows = await cur.fetchall()

    output = io.StringIO()
    writer = csv.writer(output, delimiter=";", lineterminator="\n")
    writer.writerow(["team_number", "team_name", "team_total_points", "updated_at"])
    for r in rows:
        writer.writerow(
            [
                r["team_number"],
                r["team_name"],
                r["team_total_points"],
                r["updated_at"] or "",
            ]
        )

    filename = f"rating_teams_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
    path = f"/tmp/{filename}"
    with open(path, "wb") as f:
        f.write(output.getvalue().encode("utf-8-sig"))

    await message.answer_document(FSInputFile(path, filename=filename))

async def _export_participants(message: Message):
    async with aiosqlite.connect(DB_PATH) as db:
        db.row_factory = aiosqlite.Row
        await db.execute("PRAGMA foreign_keys=ON;")
        cur = await db.execute(
            """
            SELECT user_id, fio, team_number, role, badge_number, reiting, balance, date_registered
            FROM users
            ORDER BY team_number, fio
            """
        )
        rows = await cur.fetchall()

    output = io.StringIO()
    writer = csv.writer(output, delimiter=";", lineterminator="\n")
    writer.writerow(["tg_id", "fio", "team_number", "role", "badge_number", "reiting", "balance", "date_registered"])
    for r in rows:
        writer.writerow(
            [
                r["tg_id"],
                r["fio"] or "",
                r["team_number"] if r["team_number"] is not None else "",
                r["role"] or "",
                r["badge_number"] if r["badge_number"] is not None else "",
                r["reiting"],
                r["balance"],
                r["date_registered"] or "",
            ]
        )

    filename = f"participants_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
    path = f"/tmp/{filename}"
    with open(path, "wb") as f:
        f.write(output.getvalue().encode("utf-8-sig"))

    await message.answer_document(FSInputFile(path, filename=filename))

async def _export_logs(message: Message):
    async with aiosqlite.connect(DB_PATH) as db:
        db.row_factory = aiosqlite.Row
        await db.execute("PRAGMA foreign_keys=ON;")
        cur = await db.execute(
            """
            SELECT id, event, actor_user_id, adresat_user_id, badge_number, role,
                   complaint_id, file_row_id, tg_file_id, solution, created_at
            FROM audit_log
            ORDER BY id
            """
        )
        rows = await cur.fetchall()

    output = io.StringIO()
    writer = csv.writer(output, delimiter=";", lineterminator="\n")
    writer.writerow([
        "id",
        "event",
        "actor_user_id",
        "adresat_user_id",
        "badge_number",
        "role",
        "complaint_id",
        "file_row_id",
        "tg_file_id",
        "solution",
        "created_at",
    ])
    for r in rows:
        writer.writerow(
            [
                r["id"],
                r["event"],
                r["actor_user_id"] if r["actor_user_id"] is not None else "",
                r["adresat_user_id"] if r["adresat_user_id"] is not None else "",
                r["badge_number"] if r["badge_number"] is not None else "",
                r["role"] or "",
                r["complaint_id"] if r["complaint_id"] is not None else "",
                r["file_row_id"] if r["file_row_id"] is not None else "",
                r["tg_file_id"] or "",
                r["solution"] or "",
                r["created_at"] or "",
            ]
        )

    filename = f"audit_log_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
    path = f"/tmp/{filename}"
    with open(path, "wb") as f:
        f.write(output.getvalue().encode("utf-8-sig"))

    await message.answer_document(FSInputFile(path, filename=filename))

async def start_handler(message: Message, state: FSMContext):
    user_id = message.from_user.id
    active = await get_active_users()
    if not active_sessions.get(user_id) and user_id not in active:
        registration[user_id] = User(tg_id=user_id, username=message.from_user.username)
        await message.answer(
            "Приветствую! Для начала надо пройти регистрацию.\n"
            "Введите ваш номер бейджа.",
        )
        await state.set_state(Reg.waiting_for_bage_number)
    else:
        user = await get_user(user_id)
        await log_login(user.tg_id, user.badge_number, user.role)
        active_sessions[user_id] = user
        await show_main_menu(message.bot, user_id, state=state)

async def send_files(bot: Bot, complaint_id: int, user_id: int = None) -> str:
    complaint = await get_complaint(complaint_id)
    if not complaint:
        return "Жалоба не найдена."

    user_id = user_id
    files = await get_files_by_complaint(complaint_id)
    if not files:
        return "Файлов в жалобе нет."
    
    for file in files:
        try:
            if 'image' in file.mime_type:
                await bot.send_photo(chat_id=user_id, photo=file.tg_file_id, caption=f"Файл из жалобы: {file.file_name}")
            elif 'video' in file.mime_type:
                await bot.send_video(chat_id=user_id, video=file.tg_file_id, caption=f"Файл из жалобы: {file.file_name}")
            else:
                await bot.send_document(chat_id=user_id, document=file.tg_file_id, caption=f"Файл из жалобы: {file.file_name}")
        except Exception as e:
            print(f"Error sending file {file.tg_file_id}: {e}")
    return f"Файлы отправлены пользователю {user_id}."

async def show_main_menu(bot: Bot, user_id: int, state: FSMContext | None = None,): 
    role = active_sessions[user_id].role
    match role:
        case "Участник":
            await bot.send_message(
                chat_id=user_id,
                text="Главное меню.",
                reply_markup=get_main_menu_student_keyboard(),
            )
            if state:
                await state.set_state(MainMenu.main_menu_student)

        case "Организатор":
            await bot.send_message(
                chat_id=user_id,
                text="Главное меню.",
                reply_markup=get_main_menu_organizer_keyboard(),
            )
            if state:
                await state.set_state(MainMenu.main_menu_organizer)

        case "РПГ":
            await bot.send_message(
                chat_id=user_id,
                text="Главное меню.",
                reply_markup=get_main_menu_rpg_organizer_keyboard(),
            )
            if state:
                await state.set_state(MainMenu.main_menu_rpg_organizer)

        case "Администратор":
            await bot.send_message(
                chat_id=user_id,
                text="Главное меню.",
                reply_markup=get_main_menu_admins_keyboard(),
            )
            if state:
                await state.set_state(MainMenu.main_menu_admins)

        case "Рейтинг":
            await bot.send_message(
                chat_id=user_id,
                text="Главное меню.",
                reply_markup=get_main_menu_rating_team_keyboard(),
            )
            if state:
                await state.set_state(MainMenu.main_menu_rating_team)

        case "Медиа":
            await bot.send_message(
                chat_id=user_id,
                text="Главное меню.",
                reply_markup=get_main_menu_media_team_keyboard(),
            )
            if state:
                await state.set_state(MainMenu.main_menu_media)

async def send_complaint_files(bot: Bot, chat_id: int, complaint_id: int):
    rows = await get_files_by_complaint_id(complaint_id)

    if not rows:
        await bot.send_message(chat_id, "Файлы не прикреплены.")
        return

    for r in rows:
        tg_file_id = r["tg_file_id"]
        mime = (r["mime_type"] or "").lower()

        if mime.startswith("video"):
            await bot.send_video(chat_id, tg_file_id)
        else:
            # по умолчанию считаем фото
            await bot.send_photo(chat_id, tg_file_id)

async def notify_all_reiting_team(bot: Bot, complaint: Complaint, state: FSMContext):
    team = await get_raiting_team_tg()
    for member_id in team:
        if member_id < 1000:
            continue
        await bot.send_message(
            member_id,
            "Пришла срочная жалоба! Ответить на неё?",
            reply_markup=get_yes_no_keyboard()
        )
        alarm.setdefault(member_id, [])
        alarm[member_id].append(complaint.complaint_id)

async def notify_persone_room_problems(bot: Bot, complaint: Complaint, desicion: str, state: FSMContext):
    text = 'удовлетворительное решение' if desicion == 'agree' else 'неудовлетворительное решение'
    await bot.send_message(complaint.user_id, 'Ваша жалоба на комнату была рассмотрена и принято ' + text)
    await bot.send_message(complaint.user_id, f'Жалоба: \n{complaint.description}')
    await show_main_menu(bot, complaint.user_id, state)

async def send_complaint_notify(bot: Bot, complaint: Complaint, state: FSMContext):
    fr = await get_user(complaint.user_id)
    adr = await get_user_by_badge(complaint.adresat)
    target_tg_id = adr.tg_id if adr else None

    if complaint.complaint_id and target_tg_id is not None:
        rows = await get_files_by_complaint_id(complaint.complaint_id)
        for r in rows:
            tg_file_id = r["tg_file_id"]
            mime = (r["mime_type"] or "").lower()
            if mime.startswith("video"):
                await bot.send_video(target_tg_id, tg_file_id)
            else:
                await bot.send_photo(target_tg_id, tg_file_id)
    
    fine = violetion_vines.get(complaint.violetion, 0)
    if adr is None:
        return
    
    if adr.role == 'Участник':
        if fr.role != 'Участник':
            if target_tg_id is not None:
                await bot.send_message(
                    target_tg_id,
                    "На вас пришла срочная жалоба от организатора.\n"
                    f"Время жалобы: {complaint.date_created}.\n"
                    f"Категория жалобы: {complaint.violetion}.\n"
                    f"Описание: {complaint.description}",
                )
                await update_reiting(adr.badge_number, fine)
                await update_execution(complaint.complaint_id, 'done')
        else:
            if target_tg_id is not None:
                await bot.send_message(
                    target_tg_id,
                    "На вас пришла срочная жалоба от участника.\n"
                    f"Время жалобы: {complaint.date_created}.\n"
                    f"Категория жалобы: {complaint.violetion}.\n"
                    f"Описание: {complaint.description}",
                )

    else:
        if fr.role != 'Участник':
            if target_tg_id is not None:
                await bot.send_message(
                    target_tg_id,
                    "На вас пришла срочная жалоба от организатора.\n"
                    f"Время жалобы: {complaint.date_created}.\n"
                    f"Категория жалобы: {complaint.violetion}.\n"
                    f"Описание: {complaint.description}",
                )
        else:
            if target_tg_id is not None:
                await bot.send_message(
                    target_tg_id,
                    "На вас пришла срочная жалоба от участника.\n"
                    f"Время жалобы: {complaint.date_created}.\n"
                    f"Категория жалобы: {complaint.violetion}.\n"
                    f"Описание: {complaint.description}",
                )
        await notify_all_reiting_team(bot, complaint, state)
    
async def send_complaint_notify_soon(bot: Bot, complaint: Complaint, state: FSMContext):
    fr = await get_user(complaint.user_id)
    adr = await get_user_by_badge(complaint.adresat)
    target_tg_id = adr.tg_id if adr else None

    if complaint.complaint_id and target_tg_id is not None:
        rows = await get_files_by_complaint_id(complaint.complaint_id)
        for r in rows:
            tg_file_id = r["tg_file_id"]
            mime = (r["mime_type"] or "").lower()
            if mime.startswith("video"):
                await bot.send_video(target_tg_id, tg_file_id)
            else:
                await bot.send_photo(target_tg_id, tg_file_id)
    
    fine = violetion_vines.get(complaint.violetion, 0)
    if adr is None:
        return
    
    if adr.role == 'Участник':
        if fr.role != 'Участник':
            if target_tg_id is not None:
                await bot.send_message(
                    target_tg_id,
                    "На вас пришла жалоба от организатора.\n"
                    f"Время жалобы: {complaint.date_created}.\n"
                    f"Категория жалобы: {complaint.violetion}.\n"
                    f"Описание: {complaint.description}",
                )
                await subtract_rating(adr.badge_number, fine)
                await update_execution(complaint.complaint_id, 'done')
        else:
            if target_tg_id is not None:
                req_id = await create_point_request(
                    target_badge=adr.badge_number,
                    points=fine,
                    reason="Несоблюдение правил"
                )
                await bot.send_message(
                    target_tg_id,
                    "На вас пришла жалоба от участника.\n"
                    f"Время жалобы: {complaint.date_created}.\n"
                    f"Категория жалобы: {complaint.violetion}.\n"
                    f"Описание: {complaint.description}",
                    reply_markup=decision_kb(req_id)
                )

    else:
        if fr.role != 'Участник':
            if target_tg_id is not None:
                await bot.send_message(
                    target_tg_id,
                    "На вас пришла срочная жалоба от организатора.\n"
                    f"Время жалобы: {complaint.date_created}.\n"
                    f"Категория жалобы: {complaint.violetion}.\n"
                    f"Описание: {complaint.description}",
                )
        else:
            if target_tg_id is not None:
                await bot.send_message(
                    target_tg_id,
                    "На вас пришла срочная жалоба от участника.\n"
                    f"Время жалобы: {complaint.date_created}.\n"
                    f"Категория жалобы: {complaint.violetion}.\n"
                    f"Описание: {complaint.description}",
                )
        await notify_all_reiting_team(bot, complaint, state)

async def notify_user_reiting(bot: Bot, choice: str, badge_number: int, amount: int):
    user = await get_user_by_badge(badge_number)
    if user.tg_id:
        if user.tg_id in active_sessions:
            match choice:
                case 'add':
                    await bot.send_message(user.tg_id, f'Вам начислили {amount} единц рейтинга')
                case 'subtract':
                    await bot.send_message(user.tg_id, f'С вас было снято {amount} единц рейтинга')

async def notify_rpg_buy(bot: Bot, user: User, product: Product):
    rpg_users = await get_rpg_users()

    text = (
        "Покупка в магазине.\n"
        f"Товар: {product.name}\n"
        f"Получатель: {user.fio} | {user.badge_number}"
    )

    for rpg in rpg_users:
        if rpg.tg_id and rpg.tg_id in active_sessions:
            await bot.send_message(
                chat_id=rpg.tg_id,
                text=text,
            )

async def send_complaint_room_problems(bot: Bot, complaint: Complaint):
    admins = await get_admins()
    author = await get_user(complaint.user_id)

    text = (
        f"Статус: {complaint.status}\n"
        f"Дата: {complaint.date_created}\n"
        f"От: {author.fio} @{author.username}\n"
        f"Жалоба: {complaint.description}\n"
    )

    for admin in admins:
        if admin.tg_id and admin.tg_id in active_sessions:
            await send_files(bot, complaint.complaint_id, admin.tg_id)
            await bot.send_message(
                chat_id=admin.tg_id,
                text=text,
                reply_markup=get_agree_disagree_keyboard(complaint.complaint_id),
            )

async def show_next_room_problem(bot: Bot, user_id: int, state: FSMContext):
    complaints = await get_room_problems()  # должно возвращать только "open"/необработанные
    if not complaints:
        await bot.send_message(user_id, "Нет новых комнатных жалоб.")
        await show_main_menu(bot, user_id, state)
        return

    complaint = complaints[0]

    await send_files(bot, complaint.complaint_id, user_id)

    author = await get_user(complaint.user_id)
    text = (
        f"Статус: {complaint.status}\n"
        f"Дата: {complaint.date_created}\n"
        f"От: {author.fio} @{author.username}\n"
        f"Жалоба: {complaint.description}\n"
    )

    await bot.send_message(
        chat_id=user_id,
        text=text,
        reply_markup=get_agree_disagree_keyboard(complaint.complaint_id),
    )
    await state.set_state(MainMenu.manage_rooms)

async def main():
    global bot_instance
    bot_instance = Bot(token=API_TOKEN)
    dp = Dispatcher(storage=MemoryStorage())

    dp.message.register(start_handler, CommandStart())
    dp.include_router(router)

    await load_datastore()
    actives = await get_active_users()
    for i in actives:
        active_sessions[i] = await get_user(i)
    await dp.start_polling(bot_instance)

if __name__ == "__main__":
    asyncio.run(main())
